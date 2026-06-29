#!/usr/bin/env python3
"""
Generate Kyle's local CFO budget/spend decision dashboard.

Read-only against OneDrive source workbooks. Writes derived HTML/JSON outputs
inside the Hermes-CFO workspace only.
"""
from __future__ import annotations

import json
import math
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from cfo_stripe_theme import apply_stripe_theme, ensure_theme_file

ONEDRIVE = Path("/Users/snswcommunications/Library/CloudStorage/OneDrive-Seventh-dayAdventistChurch(SouthPacific)")
WORKSPACE = Path("/Users/snswcommunications/Hermes-CFO")
OUT_DIR = WORKSPACE / "briefings" / "dashboards"
OUT_HTML = OUT_DIR / "cfo-budget-decision-dashboard.html"
OUT_JSON = OUT_DIR / "cfo-budget-decision-dashboard-data.json"
OUT_HEALTH = OUT_DIR / "cfo-budget-decision-dashboard-health.json"

BUDGET_FILE = ONEDRIVE / "Files - SNSW-Finance - Finance/1. SNC/Budget/2026/Budgets 2026.xlsx"
APPROVED_BUDGET_PDF = ONEDRIVE / "Files - SNSW-Finance - Finance/1. SNC/Budget/2026/Final budget 2026.pdf"
APPROVED_BUDGET_EMAIL_NOTE = WORKSPACE / "email-knowledge/01-mail/all/2026-06-18 - Re Evangelism Fund - 62d38be829c805ec.md"
OPERATING_DIR = ONEDRIVE / "Files - SNSW-Finance - Finance/1. SNC/Operating Statements/2026"
SUMMARY_FILE = OPERATING_DIR / "Operating statement graphs May 2026- IN PROGRESS.xlsx"

APPROVED_TOTALS = {
    "income": 8032932.0,  # 5,036,632 conference income/appropriations + 2,996,300 AAV income
    "expense": 7896544.0, # 5,638,117 conference expenditure + 2,258,427 AAV expenditure
    "net": 136388.0,
}
APPROVED_TOP_EXPENSE_BUDGET = [
    ("Field Expense", 3177120.0),
    ("Adventist Alpine Village Expenditure", 2258427.0),
    ("Administration & General Expenses", 1549196.0),
    ("Departmental Activities", 554823.0),
    ("Annual Convention Expense", 193620.0),
    ("Faith FM Expenses", 82557.0),
    ("Evangelism / Pastoral & Lay Outreach", 62000.0),
    ("Conference House Expenses", 11300.0),
    ("Miscellaneous Expense", 7500.0),
    ("Appropriations Paid", 3000.0),
]
APPROVED_TOP_INCOME_BUDGET = [
    ("Tithe available for use", 4086524.0),
    ("Adventist Alpine Village Income", 2996300.0),
    ("Education System Contribution", 511608.0),
    ("Appropriations & Interest", 202800.0),
    ("Annual Convention Income", 87500.0),
    ("Bible Worker Fund", 50000.0),
    ("Sundry Income", 35000.0),
    ("Conference House Rents", 20800.0),
]
APPROVED_FUNCTION_BUDGETS = {
    "FIELD": 3177120.0,
    "ADVENTIST ALPINE VILLAGE": 2258427.0,
    "ADMINISTRATION": 1549196.0,
    "YOUTH MINISTRY": 274288.0,
    "BIG CAMP": 193620.0,
    "MINISTERIAL": 128586.0,
    "COMMUNICATIONS": 99200.0,
    "FAITH FM ADMINISTRATION": 82557.0,
    "FAITH FM": 82557.0,
    "EVANGELISM": 62000.0,
    "PERSONAL MINISTRIES": 52750.0,
    "PROPERTIES": 11300.0,
    "OTHER OPERATIONS": 7500.0,
}
APPROVED_LANE_BUDGETS = {
    "evangelism": 62000.0,
    "faith_fm": 82557.0,
    "youth": 274288.0,
    "president_discretionary": 20000.0,
}

DECISION_LANES = [
    {
        "id": "evangelism",
        "title": "Evangelism request",
        "question": "An evangelism budget request came in. Can we afford it?",
        "function_terms": ["EVANGELISM"],
        "detail_terms": ["evangelism", "bible worker", "outreach", "atsim field evangelism"],
        "exclude_terms": ["income"],
        "default_request": 5000,
        "owner_hint": "Evangelism / outreach lane",
    },
    {
        "id": "president_discretionary",
        "title": "President discretionary",
        "question": "Justin was invited to the USA. Can the President afford it?",
        "function_terms": ["ADMINISTRATION"],
        "detail_terms": ["president"],
        "exclude_terms": ["income"],
        "default_request": 3500,
        "owner_hint": "President / administration discretionary lane",
    },
    {
        "id": "faith_fm",
        "title": "Faith FM / studio equipment",
        "question": "Faith FM needs microphones. Can we afford it?",
        "function_terms": ["FAITH FM", "EVANGELISM"],
        "detail_terms": ["faith fm", "radio", "canberra radio station", "media"],
        "exclude_terms": ["income", "sale of goods"],
        "default_request": 2500,
        "owner_hint": "Faith FM / radio ministry lane",
    },
    {
        "id": "youth",
        "title": "Youth ministry",
        "question": "Can Youth absorb this request?",
        "function_terms": ["YOUTH"],
        "detail_terms": ["youth"],
        "exclude_terms": ["income"],
        "default_request": 3000,
        "owner_hint": "Youth ministry lane",
    },
]


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def mtime_iso(path: Path) -> str | None:
    if not path.exists():
        return None
    return datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="minutes")


def money_num(value: Any) -> float:
    try:
        if value is None or value == "":
            return 0.0
        if isinstance(value, str):
            value = value.replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def fmt_money(value: Any) -> str:
    x = money_num(value)
    s = f"${abs(x):,.0f}"
    return f"({s})" if x < 0 else s


def safe_text(v: Any) -> str:
    return "" if v is None else str(v).strip()


def classify_account(acct: str) -> str:
    if acct.startswith("6"):
        return "Income"
    if acct.startswith("7"):
        return "Expense"
    if acct.startswith("8"):
        return "Trading"
    if acct.startswith("9"):
        return "Clearing/Other"
    return "Other"


def load_budget() -> dict[str, Any]:
    # Approved source: Final budget 2026.pdf, presented to Board on 15 Feb 2026.
    # The old Budgets 2026.xlsx import is deliberately not used here because it does not tie to the final PDF.
    annual = dict(APPROVED_TOTALS)
    p5 = {"income": annual["income"] / 12, "expense": annual["expense"] / 12}
    p5["net"] = p5["income"] - p5["expense"]
    jan_may = {"income": annual["income"] * 5 / 12, "expense": annual["expense"] * 5 / 12}
    jan_may["net"] = jan_may["income"] - jan_may["expense"]
    return {
        "source": str(APPROVED_BUDGET_PDF),
        "approval_email_note": str(APPROVED_BUDGET_EMAIL_NOTE),
        "modified": mtime_iso(APPROVED_BUDGET_PDF),
        "row_count": None,
        "basis_note": "Approved annual totals from Final budget 2026.pdf. Period 05 and Jan-May references are elapsed-year pace checks, not monthly-budget precision.",
        "annual": annual,
        "period_05": p5,
        "jan_may": jan_may,
        "top_expense_budget": APPROVED_TOP_EXPENSE_BUDGET,
        "top_income_budget": APPROVED_TOP_INCOME_BUDGET,
    }


def find_detailed_workbook() -> tuple[Path | None, list[str]]:
    warnings = []
    candidates = []
    for p in OPERATING_DIR.glob("*Velixo*.xls*"):
        try:
            wb = load_workbook(p, read_only=True, data_only=True, keep_links=False)
            if "Rpt B-Functions " not in wb.sheetnames:
                continue
            ws = wb["Rpt B-Functions "]
            score = 0
            nonzero_budget = 0
            nonzero_actual = 0
            for row in ws.iter_rows(min_row=11, values_only=True):
                budget = abs(money_num(row[4] if len(row) > 4 else 0))
                actual = abs(money_num(row[5] if len(row) > 5 else 0))
                if budget:
                    nonzero_budget += 1
                if actual:
                    nonzero_actual += 1
            score = nonzero_budget * 2 + nonzero_actual
            candidates.append((score, p.stat().st_mtime, p, nonzero_budget, nonzero_actual))
        except Exception as e:
            warnings.append(f"Could not inspect detailed workbook {p.name}: {e}")
    if not candidates:
        return None, warnings + ["No detailed Velixo workbook with Rpt B-Functions found."]
    candidates.sort(reverse=True)
    best = candidates[0]
    warnings.append(f"Detailed source selected by non-zero budget/actual score: {best[2].name} ({best[3]} budget rows, {best[4]} actual rows).")
    return best[2], warnings


def load_summary() -> tuple[dict[str, Any], list[str]]:
    warnings = []
    if not SUMMARY_FILE.exists():
        return {}, [f"Summary workbook missing: {SUMMARY_FILE}"]
    wb = load_workbook(SUMMARY_FILE, read_only=True, data_only=True, keep_links=False)
    ws = wb["Dashboard"]
    actual = {
        "conference_income": money_num(ws["B2"].value),
        "conference_expense": money_num(ws["B3"].value),
        "conference_net": money_num(ws["B4"].value),
        "aav_income": money_num(ws["D2"].value),
        "aav_expense": money_num(ws["D3"].value),
        "aav_net": money_num(ws["D4"].value),
        "overall_income": money_num(ws["F2"].value),
        "overall_expense": money_num(ws["F3"].value),
        "overall_net": money_num(ws["F4"].value),
    }
    cash_rows = []
    for row in range(10, 15):
        vals = [ws.cell(row=row, column=c).value for c in range(1, 8)]
        if vals[0]:
            cash_rows.append({"account": vals[0], "type": vals[1], "may": money_num(vals[6])})
    warnings.append("Summary workbook top figures do not explicitly label month-only vs YTD; dashboard shows budget reference points rather than forcing a basis.")
    return {"source": str(SUMMARY_FILE), "modified": mtime_iso(SUMMARY_FILE), "actual": actual, "cash_rows": cash_rows}, warnings


def parse_detailed(path: Path | None) -> tuple[dict[str, Any], list[str]]:
    if not path:
        return {"source": None, "functions": [], "lines": []}, ["No detailed source available."]
    warnings = []
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    ws = wb["Rpt B-Functions "]
    current_function = None
    lines = []
    functions = defaultdict(lambda: {"budget": 0.0, "actual": 0.0, "income_budget": 0.0, "income_actual": 0.0, "expense_budget": 0.0, "expense_actual": 0.0, "line_count": 0})
    for row in ws.iter_rows(min_row=11, values_only=True):
        a, b, c, d = [safe_text(row[i]) if len(row) > i else "" for i in range(4)]
        budget = money_num(row[4] if len(row) > 4 else 0)
        actual = money_num(row[5] if len(row) > 5 else 0)
        var = money_num(row[6] if len(row) > 6 else 0)
        # Function heading: only column A has text and not a total/surplus line.
        if a and not b and not c and not d and "Surplus/Deficit" not in a:
            current_function = a
            continue
        if "Surplus/Deficit" in a:
            # summary rows are useful but not line-level spend decisions.
            fname = a.replace(" Surplus/Deficit", "")
            functions[fname]["budget"] = budget
            functions[fname]["actual"] = actual
            continue
        if not current_function or not d:
            continue
        if budget == 0 and actual == 0:
            continue
        lower = d.lower()
        is_income = "income" in lower or "sale of goods" in lower or budget > 0
        is_expense = not is_income
        item = {"function": current_function, "line": d, "budget": budget, "actual": actual, "variance": var, "is_income": is_income, "is_expense": is_expense}
        lines.append(item)
        f = functions[current_function]
        f["line_count"] += 1
        if is_income:
            f["income_budget"] += budget
            f["income_actual"] += actual
        else:
            f["expense_budget"] += abs(budget)
            f["expense_actual"] += abs(actual)
    function_rows = []
    for name, vals in functions.items():
        approved_budget = APPROVED_FUNCTION_BUDGETS.get(name.upper())
        vals["source_expense_budget"] = vals["expense_budget"]
        if approved_budget is not None:
            vals["expense_budget"] = approved_budget
        else:
            # Do not carry forward old Budgets 2026.xlsx function budgets for functions absent from the approved PDF.
            vals["expense_budget"] = 0.0
        expense_remaining = vals["expense_budget"] - vals["expense_actual"]
        used_pct = (vals["expense_actual"] / vals["expense_budget"] * 100) if vals["expense_budget"] else None
        function_rows.append({"name": name, **vals, "expense_remaining": expense_remaining, "used_pct": used_pct})
    function_rows.sort(key=lambda x: abs(x.get("expense_budget", 0)), reverse=True)
    return {"source": str(path), "modified": mtime_iso(path), "functions": function_rows, "lines": lines}, warnings


def build_decision_cards(detail: dict[str, Any]) -> list[dict[str, Any]]:
    lines = detail.get("lines", [])
    cards = []
    for lane in DECISION_LANES:
        matches = []
        for line in lines:
            hay = f'{line["function"]} {line["line"]}'.lower()
            function_ok = any(t.lower() in line["function"].lower() for t in lane["function_terms"])
            detail_ok = any(t.lower() in hay for t in lane["detail_terms"])
            excluded = any(t.lower() in hay for t in lane["exclude_terms"])
            if function_ok and detail_ok and not excluded and line["is_expense"]:
                matches.append(line)
        budget = APPROVED_LANE_BUDGETS.get(lane["id"], sum(abs(x["budget"]) for x in matches))
        actual = sum(abs(x["actual"]) for x in matches)
        remaining = budget - actual
        request = lane["default_request"]
        after = remaining - request
        if budget <= 0:
            status = "Check source"
            status_class = "warn"
            advice = "No clear expense budget found for this lane. Treat as CFO review required."
        elif after >= max(1000, budget * 0.10):
            status = "Likely affordable"
            status_class = "good"
            advice = f"A {fmt_money(request)} request still leaves about {fmt_money(after)}."
        elif after >= 0:
            status = "Possible, but tight"
            status_class = "warn"
            advice = f"A {fmt_money(request)} request fits but leaves only {fmt_money(after)}."
        else:
            status = "Not affordable in lane"
            status_class = "bad"
            advice = f"A {fmt_money(request)} request would exceed the visible lane by {fmt_money(abs(after))}."
        cards.append({**lane, "budget": budget, "actual": actual, "remaining": remaining, "used_pct": (actual / budget * 100 if budget else None), "example_request": request, "after_request": after, "status": status, "status_class": status_class, "advice": advice, "matched_lines": matches[:8], "match_count": len(matches)})
    return cards


def html_escape(x: Any) -> str:
    import html
    return html.escape("" if x is None else str(x))


def render_dashboard(data: dict[str, Any]) -> str:
    budget = data["budget"]
    summary = data["summary"]
    actual = summary.get("actual", {})
    detail = data["detail"]
    cards = data["decision_cards"]
    health = data["health"]
    top_functions = [f for f in detail.get("functions", []) if f.get("expense_budget", 0) > 0][:14]
    actual_period = "May 2026 operating dashboard (source does not prove month-only vs YTD)"
    budget_period = "Approved FY2026 budget: Final budget 2026.pdf. Period 05 and Jan-May are elapsed-year pace checks, not monthly-budget precision."
    detail_period = "Detailed Velixo operating report selected from 2026 operating statements folder; budget columns overridden to approved PDF control totals where available"
    chart_payload = json.dumps({
        "p05": [round(budget["period_05"]["income"]), round(budget["period_05"]["expense"])],
        "janmay": [round(budget["jan_may"]["income"]), round(budget["jan_may"]["expense"])],
        "actual": [round(actual.get("overall_income", 0)), round(actual.get("overall_expense", 0))],
    })
    def card_html(c: dict[str, Any]) -> str:
        pct = "—" if c["used_pct"] is None else f'{c["used_pct"]:.0f}% used'
        lines = "".join(f"<li>{html_escape(x['function'])}: {html_escape(x['line'])} — budget {fmt_money(abs(x['budget']))}, spend {fmt_money(abs(x['actual']))}</li>" for x in c["matched_lines"][:4]) or "<li>No clear matching spend lines found.</li>"
        return f"""
        <section class="card decision {c['status_class']}">
          <div class="row"><div><h3>{html_escape(c['title'])}</h3></div><span class="pill {c['status_class']}">{html_escape(c['status'])}</span></div>
          <div class="metrics three"><div><b>{fmt_money(c['budget'])}</b><span>FY2026 visible budget lines</span></div><div><b>{fmt_money(c['actual'])}</b><span>Spend in selected 2026 Velixo report</span></div><div><b>{fmt_money(c['remaining'])}</b><span>Remaining vs FY2026 visible budget</span></div></div>
          <div class="bar"><i style="width:{min(100, c['used_pct'] or 0):.0f}%"></i></div><div class="small">{pct}. Example request: {fmt_money(c['example_request'])}. {html_escape(c['advice'])}</div>
          <details><summary>Matched source lines ({c['match_count']})</summary><ul>{lines}</ul></details>
        </section>"""
    warning_html = "".join(f"<li>{html_escape(w)}</li>" for w in health["warnings"])
    function_bits = []
    for f in top_functions:
        rem_class = "bad" if f["expense_remaining"] < 0 else "good"
        used = "—" if f["used_pct"] is None else f"{f['used_pct']:.0f}%"
        function_bits.append(
            f"<tr><td>{html_escape(f['name'])}</td><td>{fmt_money(f['expense_budget'])}</td>"
            f"<td>{fmt_money(f['expense_actual'])}</td><td class='{rem_class}'>{fmt_money(f['expense_remaining'])}</td><td>{used}</td></tr>"
        )
    function_rows = "".join(function_bits)
    pressure = sorted(top_functions, key=lambda f: f.get("expense_remaining", 0))[:8]
    pressure_bits = []
    for f in pressure:
        pressure_class = "bad" if f["expense_remaining"] < 0 else "good"
        pressure_used = "—" if f.get("used_pct") is None else f"{f['used_pct']:.0f}%"
        pressure_bits.append(
            f"<tr><td>{html_escape(f['name'])}</td><td>{fmt_money(f['expense_budget'])}</td>"
            f"<td>{fmt_money(f['expense_actual'])}</td><td class='{pressure_class}'>{fmt_money(f['expense_remaining'])}</td>"
            f"<td>{pressure_used}</td></tr>"
        )
    pressure_rows = "".join(pressure_bits)
    over_budget = [f for f in detail.get("functions", []) if f.get("expense_budget", 0) > 0 and f.get("expense_remaining", 0) < 0]
    tight_budget = [f for f in detail.get("functions", []) if f.get("expense_budget", 0) > 0 and (f.get("used_pct") or 0) >= 85]
    largest_overrun = min(over_budget, key=lambda f: f.get("expense_remaining", 0), default=None)
    cash_total = sum(money_num(r.get("may")) for r in summary.get("cash_rows", []))
    cash_rows = "".join(f"<tr><td>{html_escape(r['account'])}</td><td>{html_escape(r['type'])}</td><td>{fmt_money(r['may'])}</td></tr>" for r in summary.get("cash_rows", []))
    return f"""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>CFO Budget Decision Dashboard</title>
<style>
:root{{--bg:#07101f;--panel:#111c30;--panel2:#16243d;--text:#e8eefb;--muted:#9fb0cc;--line:#263956;--good:#34d399;--bad:#fb7185;--warn:#fbbf24;--accent:#67e8f9;--purple:#a78bfa}}*{{box-sizing:border-box}}body{{margin:0;background:linear-gradient(135deg,#07101f,#101827);color:var(--text);font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Arial,sans-serif}}.wrap{{max-width:1440px;margin:0 auto;padding:28px}}h1{{font-size:32px;margin:0 0 6px}}h2{{margin:0 0 12px;font-size:18px}}h3{{margin:6px 0 12px;font-size:22px}}.sub,.small{{color:var(--muted);font-size:13px;line-height:1.45}}.grid{{display:grid;grid-template-columns:repeat(12,1fr);gap:16px}}.card{{background:rgba(17,28,48,.95);border:1px solid var(--line);border-radius:18px;padding:18px;box-shadow:0 16px 40px rgba(0,0,0,.22)}}.span3{{grid-column:span 3}}.span4{{grid-column:span 4}}.span6{{grid-column:span 6}}.span8{{grid-column:span 8}}.span12{{grid-column:span 12}}.label{{font-size:12px;text-transform:uppercase;letter-spacing:.08em;color:var(--muted)}}.value{{font-size:32px;font-weight:850;margin:8px 0}}.good{{color:var(--good)}}.bad{{color:var(--bad)}}.warn{{color:var(--warn)}}.pill{{display:inline-block;padding:7px 10px;border-radius:999px;background:#1d2b45;font-weight:700;font-size:12px}}.pill.good{{background:rgba(52,211,153,.14)}}.pill.bad{{background:rgba(251,113,133,.14)}}.pill.warn{{background:rgba(251,191,36,.14)}}.row{{display:flex;justify-content:space-between;gap:12px;align-items:flex-start}}.metrics{{display:grid;gap:10px;margin:10px 0}}.metrics.three{{grid-template-columns:repeat(3,1fr)}}.metrics div{{background:rgba(255,255,255,.035);border:1px solid var(--line);border-radius:14px;padding:10px}}.metrics b{{display:block;font-size:22px}}.metrics span{{display:block;color:var(--muted);font-size:12px;margin-top:3px}}.bar{{height:10px;background:#263956;border-radius:99px;overflow:hidden;margin:10px 0}}.bar i{{display:block;height:100%;background:linear-gradient(90deg,var(--good),var(--warn... [truncated]
</style></head><body><div class="wrap">
<h1>SNSW CFO Operating Dashboard</h1>
<div class="sub">Generated {html_escape(data['generated_at'])}. Local OneDrive-derived operating view for observing pressure, trend, cash, and budget capacity.</div>
<div class="card span12" style="margin:16px 0"><div class="label">Period / basis</div><div class="small"><b>Actuals:</b> {html_escape(actual_period)}. <b>Budget:</b> {html_escape(budget_period)}. <b>Function drilldown:</b> {html_escape(detail_period)}.</div></div>
<div class="grid">
<section class="card span3"><div class="label">May 2026 operating income</div><div class="value">{fmt_money(actual.get('overall_income'))}</div><div class="small">{html_escape(actual_period)}</div></section>
<section class="card span3"><div class="label">May 2026 operating spend</div><div class="value">{fmt_money(actual.get('overall_expense'))}</div><div class="small">{html_escape(actual_period)}</div></section>
<section class="card span3"><div class="label">May 2026 operating net</div><div class="value {'bad' if actual.get('overall_net',0)<0 else 'good'}">{fmt_money(actual.get('overall_net'))}</div><div class="small">{html_escape(actual_period)}</div></section>
<section class="card span3"><div class="label">Data health</div><div class="value {'good' if health['status']=='OK' else 'warn'}">{html_escape(health['status'])}</div><div class="small">{len(health['warnings'])} warning(s)</div></section>

<section class="card span12"><div class="label">Observation layer</div><h2>What the dashboard is telling you</h2><div class="metrics three"><div><b>{fmt_money(actual.get('overall_net'))}</b><span>May 2026 operating net visible in summary workbook</span></div><div><b>{len(over_budget)}</b><span>Functions over FY2026 visible budget in selected Velixo report</span></div><div><b>{len(tight_budget)}</b><span>Functions at or above 85% of FY2026 visible budget</span></div></div><div class="small">Largest visible overrun: <b>{html_escape(largest_overrun['name']) if largest_overrun else 'None visible'}</b>{(' — ' + fmt_money(largest_overrun['expense_remaining'])) if largest_overrun else ''}. Visible cash across listed accounts: <b>{fmt_money(cash_total)}</b>.</div></section>

<section class="card span7"><div class="label">FY2026 budget references vs May 2026 operating actuals</div><canvas id="budgetChart"></canvas></section>
<section class="card span5"><div class="label">Pressure watchlist — {html_escape(detail_period)}</div><div class="small">Budget/spend/remaining from {html_escape(detail_period)}; health remains WARN until source month-only vs YTD basis is confirmed.</div><table><thead><tr><th>Function</th><th>Budget</th><th>Spend</th><th>Remaining</th><th>Used</th></tr></thead><tbody>{pressure_rows}</tbody></table></section>

<section class="card span6"><div class="label">Actuals from May 2026 operating dashboard</div><div class="small">Basis: {html_escape(actual_period)}</div><table><thead><tr><th>Area</th><th>Income</th><th>Spend</th><th>Net</th></tr></thead><tbody>
<tr><td>SNSW Conference</td><td>{fmt_money(actual.get('conference_income'))}</td><td>{fmt_money(actual.get('conference_expense'))}</td><td>{fmt_money(actual.get('conference_net'))}</td></tr>
<tr><td>Adventist Alpine Village</td><td>{fmt_money(actual.get('aav_income'))}</td><td>{fmt_money(actual.get('aav_expense'))}</td><td>{fmt_money(actual.get('aav_net'))}</td></tr>
<tr><td><b>Overall</b></td><td><b>{fmt_money(actual.get('overall_income'))}</b></td><td><b>{fmt_money(actual.get('overall_expense'))}</b></td><td><b>{fmt_money(actual.get('overall_net'))}</b></td></tr>
</tbody></table></section>
<section class="card span6"><div class="label">Budget reference points</div><div class="small">Basis: {html_escape(budget_period)}</div><table><thead><tr><th>Budget basis</th><th>Income</th><th>Spend</th><th>Net</th></tr></thead><tbody>
<tr><td>Period 05 only</td><td>{fmt_money(budget['period_05']['income'])}</td><td>{fmt_money(budget['period_05']['expense'])}</td><td>{fmt_money(budget['period_05']['net'])}</td></tr>
<tr><td>Periods 01–05</td><td>{fmt_money(budget['jan_may']['income'])}</td><td>{fmt_money(budget['jan_may']['expense'])}</td><td>{fmt_money(budget['jan_may']['net'])}</td></tr>
<tr><td>Full year</td><td>{fmt_money(budget['annual']['income'])}</td><td>{fmt_money(budget['annual']['expense'])}</td><td>{fmt_money(budget['annual']['net'])}</td></tr>
</tbody></table></section>
<section class="card span8"><div class="label">Largest function budgets — {html_escape(detail_period)}</div><table><thead><tr><th>Function</th><th>Budget</th><th>Spend</th><th>Remaining</th><th>Used</th></tr></thead><tbody>{function_rows}</tbody></table></section>
<section class="card span4"><div class="label">Cash balances visible — {html_escape(actual_period)}</div><table><thead><tr><th>Account</th><th>Type</th><th>{html_escape(actual_period)}</th></tr></thead><tbody>{cash_rows}</tbody></table></section>
<section class="card span12 notice"><b>How to use:</b> Start with the observation layer, then scan the pressure watchlist and largest function table. This dashboard is not for querying hypothetical examples; it is for spotting where the operating position, budget usage, cash, and source warnings require CFO judgement.</section>
<section class="card span12"><div class="label">Warnings / health</div><ul>{warning_html}</ul></section>
<section class="card span12"><div class="label">Sources</div><div class="src">Budget: {html_escape(budget['source'])}<br>Summary actuals: {html_escape(summary.get('source'))}<br>Detailed report: {html_escape(detail.get('source'))}<br>Output JSON: {html_escape(str(OUT_JSON))}<br>Health JSON: {html_escape(str(OUT_HEALTH))}</div></section>
</div></div>
<script>
const D={chart_payload};
function moneyLabel(x){{return (x<0?'($'+Math.round(Math.abs(x)).toLocaleString()+')':'$'+Math.round(x).toLocaleString());}}
function bar(id, labels, series){{const c=document.getElementById(id),ctx=c.getContext('2d');c.width=c.clientWidth*2;c.height=260*2;ctx.scale(2,2);const W=c.clientWidth,H=260,p=40,max=Math.max(1,...series.flatMap(s=>s.vals.map(v=>Math.abs(v))))*1.15;ctx.strokeStyle='#d6dff0';ctx.lineWidth=1;ctx.beginPath();ctx.moveTo(p,15);ctx.lineTo(p,H-p);ctx.lineTo(W-10,H-p);ctx.stroke();const group=(W-p-30)/labels.length,bw=group/(series.length+1);labels.forEach((lab,i)=>{{series.forEach((s,j)=>{{const val=Math.abs(s.vals[i]),x=p+20+i*group+j*(bw+4),h=val/max*(H-p-30);ctx.fillStyle=s.color;ctx.fillRect(x,H-p-h,bw,h);}});ctx.fillStyle='#64748d';ctx.font='12px Source Sans 3, system-ui';ctx.fillText(lab,p+20+i*group,H-12)}});series.forEach((s,i)=>{{ctx.fillStyle=s.color;ctx.font='12px Source Sans 3, system-ui';ctx.fillText(s.name, W-300+i*92, 20)}})}}
function lane(id){{const c=document.getElementById(id),ctx=c.getContext('2d');c.width=c.clientWidth*2;c.height=260*2;ctx.scale(2,2);const W=c.clientWidth,H=260,p=28,vals=D.cards,max=Math.max(1,...vals.map(x=>Math.abs(x.remaining)))*1.2;ctx.font='12px Source Sans 3, system-ui';vals.forEach((x,i)=>{{const y=32+i*48,w=Math.abs(x.remaining)/max*(W-170);ctx.fillStyle='#e6edf7';ctx.fillRect(135,y,W-190,20);ctx.fillStyle=x.remaining>=0?'#15be53':'#ea2261';ctx.fillRect(135,y,w,20);ctx.fillStyle='#273951';ctx.fillText(x.title,12,y+15);ctx.fillStyle=x.remaining>=0?'#108c3d':'#ea2261';ctx.fillText(moneyLabel(x.remaining),140+w+8,y+15)}})}}
bar('budgetChart',['Income','Spend'],[{{name:'P05 budget',color:'#5bd8e8',vals:D.p05}},{{name:'Jan-May',color:'#8b7cff',vals:D.janmay}},{{name:'Actual',color:'#15be53',vals:D.actual}}]);
</script></body></html>"""


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    health = {"generated_at": now_iso(), "status": "OK", "warnings": [], "errors": []}
    try:
        budget = load_budget()
    except Exception as e:
        health["errors"].append(f"Budget load failed: {e}")
        raise
    try:
        summary, warnings = load_summary()
        health["warnings"].extend(warnings)
    except Exception as e:
        health["errors"].append(f"Summary load failed: {e}")
        summary = {"source": str(SUMMARY_FILE), "modified": mtime_iso(SUMMARY_FILE), "actual": {}, "cash_rows": []}
    detailed_path, warnings = find_detailed_workbook()
    health["warnings"].extend(warnings)
    try:
        detail, warnings = parse_detailed(detailed_path)
        health["warnings"].extend(warnings)
        health["warnings"].append("Approved PDF control totals override old workbook budgets; functions not listed in the approved PDF are shown with no budget rather than carrying stale Budgets 2026.xlsx allocations.")
    except Exception as e:
        health["errors"].append(f"Detailed load failed: {e}")
        detail = {"source": str(detailed_path) if detailed_path else None, "functions": [], "lines": []}
    cards = build_decision_cards(detail)
    if health["errors"]:
        health["status"] = "ERROR"
    elif health["warnings"]:
        health["status"] = "WARN"
    data = {"generated_at": now_iso(), "budget": budget, "summary": summary, "detail": detail, "decision_cards": cards, "health": health}
    OUT_JSON.write_text(json.dumps(data, indent=2), encoding="utf-8")
    OUT_HEALTH.write_text(json.dumps(health, indent=2), encoding="utf-8")
    ensure_theme_file(OUT_DIR)
    OUT_HTML.write_text(apply_stripe_theme(render_dashboard(data)), encoding="utf-8")
    print(f"Generated {OUT_HTML}")
    print(f"Status: {health['status']}; warnings={len(health['warnings'])}; errors={len(health['errors'])}")
    return 0 if not health["errors"] else 2


if __name__ == "__main__":
    sys.exit(main())
