#!/usr/bin/env python3
from pathlib import Path
from datetime import datetime
import json, math, sys
from collections import defaultdict
from openpyxl import load_workbook
from cfo_stripe_theme import apply_stripe_theme, ensure_theme_file

ONEDRIVE=Path('/Users/snswcommunications/Library/CloudStorage/OneDrive-Seventh-dayAdventistChurch(SouthPacific)')
OUT_DIR=Path('/Users/snswcommunications/Hermes-CFO/briefings/dashboards')
OUT_HTML=OUT_DIR/'department-budget-dashboard.html'
OUT_JSON=OUT_DIR/'department-budget-dashboard-data.json'
MYOB_REPORT=OUT_DIR/'department-budget-myob-data.json'
DETAILED=ONEDRIVE/'Files - SNSW-Finance - Finance/1. SNC/Operating Statements/2026/Velixo Reports Master Adventist V1.8F.xlsx'
SUMMARY=ONEDRIVE/'Files - SNSW-Finance - Finance/1. SNC/Operating Statements/2026/Operating statement graphs May 2026- IN PROGRESS.xlsx'
BUDGET=ONEDRIVE/'Files - SNSW-Finance - Finance/1. SNC/Budget/2026/Budgets 2026.xlsx'
APPROVED_BUDGET_PDF=ONEDRIVE/'Files - SNSW-Finance - Finance/1. SNC/Budget/2026/Final budget 2026.pdf'
APPROVED_BUDGET_EMAIL_NOTE=Path('/Users/snswcommunications/Hermes-CFO/email-knowledge/01-mail/all/2026-06-18 - Re Evangelism Fund - 62d38be829c805ec.md')
APPROVED_BUDGET_BASIS='FY2026 approved annual budget — Final budget 2026.pdf, presented to Board 15 Feb 2026'

APPROVED_DEPARTMENT_BUDGETS={
    'FIELD':3177120.0,
    'ADVENTIST ALPINE VILLAGE':2258427.0,
    'ADMINISTRATION':1549196.0,
    'YOUTH MINISTRY':274288.0,
    'BIG CAMP':193620.0,
    'MINISTERIAL':128586.0,
    'COMMUNICATIONS':99200.0,
    'FAITH FM ADMINISTRATION':82557.0,
    'EVANGELISM':62000.0,
    'PERSONAL MINISTRIES / DEPARTMENT LIAISONS':52750.0,
    'PROPERTIES':11300.0,
    'OTHER OPERATIONS':7500.0,
}

APPROVED_DEPARTMENT_LINES={
    'FIELD':[
        ('Wages Taxable',1064871.0),('Fringe Benefits Budget',816423.0),('Travel & Motor Vehicle',481083.0),
        ('Superannuation - ACAST',255104.0),('Tithe Expense',234400.0),('Removal',70000.0),
        ('Book & Equipment Subsidy',56104.0),('LSL',48567.0),('Professional Development',45000.0),
        ('ADSAFE Contributions',38767.0),('Workers Compensation',35521.0),('Telephone',22080.0),
        ('Field Exp',6200.0),('Student Fees Discount',3000.0),
    ],
    'ADMINISTRATION':[
        ('Fixed Expenses total',996396.0),('Accounting Fees/Overseas services',146500.0),('Technology Expense/Software',142000.0),
        ('Travel Expense',26600.0),('President Discretionary Expenses',20000.0),('Property Usage',20000.0),
        ('Depreciation',15000.0),('General Expense',14200.0),('Legal Expenses',12000.0),('Auditing Expense',10748.0),
        ('Cleaning & Garden',9360.0),('Office Building Maintenance',4000.0),('Equipment R & M',2500.0),
        ('Professional Development',2500.0),('Church Supplies',1500.0),('Stationery',1500.0),('Telephone',1500.0),
        ('Postage & Freight',1000.0),('Student Fees Discount',550.0),('Trailer Expense',400.0),
    ],
    'EVANGELISM':[('Pastoral & Lay Outreach',62000.0)],
    'FAITH FM ADMINISTRATION':[('Faith FM fixed costs',72557.0),('Faith FM variable costs',10000.0)],
    'YOUTH MINISTRY':[('APS 10 Youth & Family Life',274288.0)],
    'MINISTERIAL':[('APS 7 Ministerial Department',128586.0)],
    'COMMUNICATIONS':[('APS 3 Communications',99200.0)],
    'PERSONAL MINISTRIES / DEPARTMENT LIAISONS':[('APS 1 Department Liaisons',52750.0)],
    'BIG CAMP':[('Annual Convention Expense',193620.0)],
    'ADVENTIST ALPINE VILLAGE':[('Adventist Alpine Village Expenditure',2258427.0)],
    'PROPERTIES':[('Conference House Expenses',11300.0)],
    'OTHER OPERATIONS':[('Miscellaneous Activities',7500.0)],
}

def n(v):
    try:
        if v is None or v=='': return 0.0
        return float(v)
    except: return 0.0

def money(v):
    v=n(v); s=f'${abs(v):,.0f}'; return f'({s})' if v<0 else s

def pct(v):
    if v is None: return '—'
    return f'{v:.0f}%'

def elapsed_ratio(label):
    text=str(label or '').lower()
    months={'jan':1,'january':1,'feb':2,'february':2,'mar':3,'march':3,'apr':4,'april':4,'may':5,'jun':6,'june':6,'jul':7,'july':7,'aug':8,'august':8,'sep':9,'sept':9,'september':9,'oct':10,'october':10,'nov':11,'november':11,'dec':12,'december':12}
    for k,v in months.items():
        if k in text:
            return v/12
    return None

def pace_label(actual, budget, ratio):
    if not ratio or not budget:
        return ('No pace basis', 0.0, None)
    expected=budget*ratio
    variance=expected-actual
    if variance >= 0:
        return (f'{money(variance)} under elapsed-year pace', variance, expected)
    return (f'{money(abs(variance))} over elapsed-year pace', variance, expected)

def esc(s):
    import html
    return html.escape('' if s is None else str(s))

def pick_source():
    # use workbook with richest budget+actual data
    candidates=[]
    for p in DETAILED.parent.glob('*Velixo*.xls*'):
        try:
            wb=load_workbook(p,read_only=True,data_only=True,keep_links=False)
            if 'Rpt B-Functions ' not in wb.sheetnames: continue
            ws=wb['Rpt B-Functions ']
            b=a=0
            for row in ws.iter_rows(min_row=11, values_only=True):
                if abs(n(row[4] if len(row)>4 else 0)): b+=1
                if abs(n(row[5] if len(row)>5 else 0)): a+=1
            candidates.append((b*2+a,p.stat().st_mtime,p,b,a))
        except Exception:
            pass
    # Prefer an actual 2026 Velixo detailed report. Some files live in the 2026 folder
    # but are cached to 2025 periods, which would make the dashboard misleading.
    detailed=[]
    for score, mtime, p, b, a in candidates:
        try:
            wb=load_workbook(p,read_only=True,data_only=True,keep_links=False)
            period=str(wb['Parameters']['B2'].value or '') if 'Parameters' in wb.sheetnames else ''
        except Exception:
            period=''
        detailed.append((('2026' in period), mtime, score, p, period))
    detailed.sort(reverse=True)
    return detailed[0][3] if detailed else DETAILED

def workbook_period_label(path):
    try:
        wb=load_workbook(path,read_only=True,data_only=True,keep_links=False)
        if 'Parameters' in wb.sheetnames:
            v=wb['Parameters']['B2'].value
            if v:
                return str(v).strip()
    except Exception:
        pass
    name=Path(path).name
    if 'May 2026' in name: return 'May 2026'
    if 'Feb 2026' in name or name.startswith('Feb '): return 'Feb 2026'
    if '2026' in name: return '2026'
    return 'period not stated in source workbook'

def source_date_label(path):
    try:
        return datetime.fromtimestamp(Path(path).stat().st_mtime).isoformat(timespec='minutes')
    except Exception:
        return None

def build_period_context(source):
    actual_period=workbook_period_label(source)
    return {
        'budget_year':'2026',
        'budget_period_label':APPROVED_BUDGET_BASIS,
        'actual_period_label':actual_period,
        'summary_period_label':'May 2026 operating summary',
        'as_of_date': actual_period,
        'budget_source_modified': source_date_label(APPROVED_BUDGET_PDF),
        'actual_source_modified': source_date_label(source),
        'summary_source_modified': source_date_label(SUMMARY),
        'period_note':'Budget figures now use the approved Final budget 2026 PDF, not the older Budgets 2026.xlsx import. Spend/actual figures are from the selected Velixo operating report period shown here.',
    }

def parse_functions(path):
    wb=load_workbook(path,read_only=True,data_only=True,keep_links=False)
    ws=wb['Rpt B-Functions ']
    current=None
    funcs=defaultdict(lambda:{'budget':0.0,'spent':0.0,'income_budget':0.0,'income_actual':0.0,'lines':[]})
    for row in ws.iter_rows(min_row=11, values_only=True):
        a=str(row[0]).strip() if row and row[0] is not None else ''
        b=str(row[1]).strip() if len(row)>1 and row[1] is not None else ''
        c=str(row[2]).strip() if len(row)>2 and row[2] is not None else ''
        d=str(row[3]).strip() if len(row)>3 and row[3] is not None else ''
        budget=n(row[4] if len(row)>4 else 0); actual=n(row[5] if len(row)>5 else 0)
        if a and not b and not c and not d and 'Surplus/Deficit' not in a:
            current=a; continue
        if not current or not d or (budget==0 and actual==0): continue
        text=d.lower()
        is_income=('income' in text or 'sale of goods' in text or budget>0)
        if is_income:
            funcs[current]['income_budget'] += budget
            funcs[current]['income_actual'] += actual
        else:
            funcs[current]['budget'] += abs(budget)
            funcs[current]['spent'] += abs(actual)
            funcs[current]['lines'].append({'line':d,'budget':abs(budget),'spent':abs(actual),'remaining':abs(budget)-abs(actual)})
    departments=[]
    for name,v in funcs.items():
        if v['budget']<=0 and v['spent']<=0: continue
        remaining=v['budget']-v['spent']
        used=(v['spent']/v['budget']*100) if v['budget'] else None
        status='over' if remaining<0 else 'tight' if used and used>85 else 'ok'
        v['lines']=sorted(v['lines'], key=lambda x: x['budget'], reverse=True)[:8]
        departments.append({'name':name,'budget':v['budget'],'spent':v['spent'],'remaining':remaining,'used_pct':used,'status':status,'income_budget':v['income_budget'],'income_actual':v['income_actual'],'lines':v['lines']})
    departments.sort(key=lambda x:x['budget'], reverse=True)
    return departments

def parse_summary():
    out={'income':0,'spend':0,'net':0,'cash':[]}
    try:
        wb=load_workbook(SUMMARY,read_only=True,data_only=True,keep_links=False)
        ws=wb['Dashboard']
        out['income']=n(ws['F2'].value); out['spend']=n(ws['F3'].value); out['net']=n(ws['F4'].value)
        for r in range(10,15):
            name=ws.cell(r,1).value
            if name: out['cash'].append({'name':name,'type':ws.cell(r,2).value,'may':n(ws.cell(r,7).value)})
    except Exception as e: out['error']=str(e)
    return out

def render(data):
    depts=data['departments']
    period=data.get('period_context',{})
    budget_period=period.get('budget_period_label','Budget period not stated')
    actual_period=period.get('actual_period_label','Actual period not stated')
    summary_period=period.get('summary_period_label','Summary period not stated')
    period_note=period.get('period_note','')
    elapsed=elapsed_ratio(actual_period)
    elapsed_text=f'{elapsed*100:.0f}% of year elapsed' if elapsed else 'elapsed-year basis unknown'
    current_month=datetime.now().month
    current_pace_ratio=current_month/12
    current_pace_label=f'{datetime.now():%B %Y} pace target ({current_pace_ratio*100:.0f}% of annual budget)'
    source_kind=period.get('source_kind')
    detail_is_live=(source_kind=='myob_live_gl_cache')
    detail_is_stale=not detail_is_live and ((actual_period or '').lower().strip() not in f'{datetime.now():%B %Y}'.lower())
    detail_status='Live MYOB actuals' if detail_is_live else ('Latest department detail is stale' if detail_is_stale else 'Latest department detail is current')
    spend_label='MYOB actual spend' if detail_is_live else 'Spend — latest indexed detail'
    detail_caveat=f'Current MYOB JournalTransaction lines through {actual_period}.' if detail_is_live else f'Spend is not current; latest line detail is {actual_period}.'
    refresh_action='Click for MYOB transaction evidence / audit trail.' if detail_is_live else 'For a June decision: refresh MYOB actuals, then this becomes Budget / Spend / Remaining as at June.'
    source_value_class='good' if detail_is_live else 'amber'
    source_value_text='MYOB' if detail_is_live else 'Stale'
    source_status_text='Live MYOB JournalTransaction actuals via Morpheus.' if detail_is_live else f'Latest department line detail indexed: {actual_period}. Refresh MYOB before decisions.'
    pace_note='Half-year benchmark for scanning FY2026 annual budgets against Jan–Jun actuals.' if detail_is_live else 'Use half-year budget as the rough June benchmark until current department actuals are indexed.'
    budget_compare_label='MYOB budget vs actual' if detail_is_live else 'Department budget authority'
    detail_source_name='MYOB transaction lines' if detail_is_live else 'selected Velixo report'
    drilldown_note='This drilldown shows MYOB transaction lines available for the stated actual period. Use the generated report pack/source manifest for audit trail; review unmapped rows as mapping/control issues before treating them as overspend.' if detail_is_live else 'This drilldown shows the spend lines available in the selected Velixo report for the stated actual period. If you need every invoice/vendor/person transaction, refresh MYOB transaction detail for that department.'
    header_action='review mapping before June decisions' if detail_is_live else 'refresh MYOB/Velixo detail before June decisions'
    budget_compare_title='Approved budget + MYOB actuals' if detail_is_live else 'Approved budget + stale detail check'
    budget_compare_note=f'Budget is approved FY2026. Actual/spend detail is live MYOB JournalTransaction data for {actual_period}.' if detail_is_live else f'Budget is approved FY2026. Actual/spend detail is sourced from MYOB JournalTransaction when live extraction is available; otherwise the page is showing legacy indexed detail ({actual_period}).'
    legend_actual='MYOB actuals' if detail_is_live else 'Stale detail'
    next_title='MYOB live report status' if detail_is_live else 'Refresh June department actuals'
    next_intro='This page is now using live MYOB JournalTransaction expense lines mapped to department/function. Review over-budget and unmapped rows as mapping/control issues, not Velixo refresh issues.' if detail_is_live else 'The May operating summary gives whole-of-entity totals, but department-level decisions need current MYOB JournalTransaction detail mapped by department/function. Until the live MYOB extractor is reachable, use the cards below as approved budget authority plus legacy detail only.'
    next_not_current_label='Needs mapping review' if detail_is_live else 'Not current'
    next_not_current_text='Over-budget/unmapped rows may reflect MYOB subaccount mapping, budget allocation, or genuine overspend; click rows for evidence.' if detail_is_live else f'Department-by-department spend detail, currently from {actual_period} unless MYOB live extraction is active.'
    next_action_text='Use the report pack evidence CSV/source manifest for audit trail, then refine department/subaccount mapping.' if detail_is_live else 'Restore Morpheus/Tailscale access, run the MYOB refresh script, and regenerate this page/report pack.'
    for d in depts:
        label, variance, expected = pace_label(d.get('spent',0), d.get('budget',0), elapsed)
        d['pace_label']=label
        d['pace_variance']=variance
        d['expected_at_elapsed']=expected
        d['current_pace_target']=d.get('budget',0)*current_pace_ratio
    visible=depts[:18]
    acs_visible=next((d for d in depts if d['name']=='ADVENTIST COMMUNITY SERVICES'), None)
    if acs_visible and all(d['name']!='ADVENTIST COMMUNITY SERVICES' for d in visible):
        visible=(visible[:-1] if len(visible)>=18 else visible) + [acs_visible]
    over=[d for d in depts if d['remaining']<0]
    tight=[d for d in depts if d['remaining']>=0 and d['used_pct'] and d['used_pct']>85]
    healthy=[d for d in depts if d['remaining']>=0 and (d['used_pct'] is None or d['used_pct']<=85)]
    cards=''.join(f"""
      <article class='dept useful-card {d['status']}' data-name='{esc(d['name']).lower()}' onclick='openDept({json.dumps(d['name'])})' title='Click to drill into spend lines'>
        <div class='dept-head'><h3>{esc(d['name'])}</h3><span>{pct(d['used_pct'])} of budget</span></div>
        <div class='decision-line'><b>Answer</b> {money(d.get('remaining',0))} {'over budget' if d.get('remaining',0)<0 else 'remaining'} · <b>Source</b> {esc(actual_period)}</div>
        <div class='budget-scale actual-scale' aria-label='{pct(d.get('used_pct'))} of annual budget used'><i style='width:{min(100,max(0,d.get('used_pct') or 0)):.1f}%'></i><span>{pct(d.get('used_pct'))} spent</span></div>
        <div class='nums useful-nums'>
          <div><b>{money(d['budget'])}</b><small>Budget</small></div>
          <div><b class='amber'>{money(d['spent'])}</b><small>{esc(spend_label)}</small></div>
          <div><b class='{('bad' if d.get('remaining',0)<0 else 'good')}'>{money(d.get('remaining',0))}</b><small>{'Over budget' if d.get('remaining',0)<0 else 'Remaining'}</small></div>
        </div>
        <div class='card-action'>{esc(refresh_action)}</div>
      </article>""" for d in visible)
    rows=''.join(f"<tr><td>{esc(d['name'])}</td><td>{money(d['budget'])}</td><td>{money(d['spent'])}</td><td>{money(d.get('expected_at_elapsed')) if d.get('expected_at_elapsed') is not None else '—'}</td><td class='{('bad' if d.get('pace_variance',0)<0 else 'good')}'>{money(d.get('pace_variance',0))}</td><td>{pct(d['used_pct'])}</td><td>{esc(elapsed_text)}</td></tr>" for d in depts)
    def compare_row(d):
        used = d.get('used_pct') or 0
        width = min(100, max(0, used))
        status = 'over' if d.get('remaining', 0) < 0 else 'tight' if used > 85 else 'ok'
        rem_label = 'Over' if d.get('remaining', 0) < 0 else 'Remaining'
        rem_class = 'bad' if d.get('remaining', 0) < 0 else 'good'
        remaining_display = money(abs(d.get('remaining',0))) if d.get('remaining',0) < 0 else money(d.get('remaining',0))
        return f"""
        <tr class='compare-row {status}' onclick='openDept({json.dumps(d['name'])})'>
          <td class='dept-name'><b>{esc(d['name'])}</b></td>
          <td class='num budget'>{money(d['budget'])}</td>
          <td class='num actual'>{money(d['spent'])}</td>
          <td class='num {rem_class}'><b>{remaining_display}</b><span>{esc(rem_label)}</span></td>
          <td class='progress-cell'><div class='progress-line'><i class='{status}' style='width:{width:.1f}%'></i></div><strong>{pct(d.get('used_pct'))}</strong></td>
          <td class='row-action'>Drilldown →</td>
        </tr>"""
    compare_rows=''.join(compare_row(d) for d in visible[:14])
    cash=''.join(f"<tr><td>{esc(c['name'])}</td><td>{esc(c['type'])}</td><td>{money(c['may'])}</td></tr>" for c in data['summary'].get('cash',[]))
    chart=json.dumps({'departments':[{'name':d['name'],'budget':round(d['budget']),'spent':round(d['spent']),'remaining':round(d['remaining']),'used':d['used_pct'] or 0,'lines':d.get('lines',[])} for d in visible], 'allDepartments':[{'name':d['name'],'budget':round(d['budget']),'spent':round(d['spent']),'remaining':round(d['remaining']),'used':d['used_pct'] or 0,'lines':d.get('lines',[])} for d in depts]})
    field = next((d for d in depts if d['name'].upper() == 'FIELD'), None)
    field_rows = ''
    if field:
        def field_row(x):
            expected=(x['budget']*elapsed) if elapsed else None
            variance=(expected-x['spent']) if expected is not None else 0.0
            return f"<tr><td>{esc(x['line'])}</td><td>{money(x['budget'])}</td><td>{money(x['spent'])}</td><td>{money(expected) if expected is not None else '—'}</td><td class='{('bad' if variance<0 else 'good')}'>{money(variance)}</td><td><input class='proj' data-field-line='{esc(x['line'])}' value='{round(x['budget'])}'></td></tr>"
        field_rows = ''.join(field_row(x) for x in field.get('lines', []))
    else:
        field_rows = f"<tr><td colspan='5'>No FIELD budget lines found in the selected Velixo report for {esc(actual_period)}.</td></tr>"
    return f"""<!doctype html><html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width, initial-scale=1'>
<title>SNSW Department Budget Dashboard</title>
<style>
:root{{--bg:#060b16;--panel:#0f1b2f;--panel2:#13243d;--text:#eef5ff;--muted:#9db0cb;--line:#27405f;--blue:#38bdf8;--violet:#a78bfa;--green:#34d399;--amber:#fbbf24;--red:#fb7185}}
*{{box-sizing:border-box}}body{{margin:0;background:radial-gradient(circle at top left,#172554,#060b16 45%),linear-gradient(135deg,#060b16,#111827);color:var(--text);font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Arial,sans-serif}}.wrap{{max-width:1500px;margin:0 auto;padding:30px}}.hero{{display:flex;justify-content:space-between;gap:24px;align-items:flex-start;margin-bottom:22px}}h1{{font-size:38px;margin:0 0 8px;letter-spacing:-.03em}}.compact-hero{{margin-bottom:14px}}.compact-hero h1{{font-size:30px}}.compact-hero .sub{{font-size:13px}}.sub{{color:var(--muted);line-height:1.45;font-size:14px}}.pill{{display:inline-flex;gap:8px;align-items:center;border:1px solid var(--line);background:rgba(15,27,47,.8);padding:10px 14px;border-radius:999px;color:var(--muted);font-size:13px}}.grid{{display:grid;grid-template-columns:repeat(12,1fr);gap:16px}}.card,.dept{{background:rgba(15,27,47,.88);border:1px solid var(--line);border-radius:22px;padding:18px;box-shadow:0 20px 55px rgba(0,0,0,.25);backdrop-filter:blur(14px)}}.span3{{grid-column:span 3}}.span4{{grid-column:span 4}}.span5{{grid-column:span 5}}.span7{{grid-column:span 7}}.span12{{grid-column:span 12}}.label{{font-size:12px;text-transform:uppercase;color:var(--muted);letter-spacing:.09em}}.value{{font-size:34px;font-weight:850;margin:8px 0}}.good{{color:var(--green)}}.bad{{color:var(--red)}}.amber{{color:var(--amber)}}.search{{width:100%;padding:14px 16px;border-radius:16px;border:1px solid var(--line);background:#091426;color:var(--text);font-size:15px;outline:none}}canvas{{width:100%;height:400px}}.dept-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}}.dept{{padding:16px}}.dept.over{{border-color:rgba(251,113,133,.65)}}.dept.tight{{border-color:rgba(251,191,36,.7)}}.dept-head{{display:flex;justify-content:space-between;gap:10px;align-items:start}}.dept h3{{margin:0;font-size:17px;letter-spacing:-.02em}}.dept-head span{{color:var(--muted);font-size:12px;white-space:nowrap}}.bar{{position:relative;height:12px;background:#25364f;border-radius:99px;overflow:hidden;margin:14px 0}}.bar .spent{{display:block;height:100%;background:linear-gradient(90deg,var(--blue),var(--violet));border-radius:99px}}.nums{{display:grid;grid-template-columns:repeat(3,1fr);gap:8px}}.nums b{{display:block;font-size:18px}}.nums small{{color:var(--muted);font-size:11px}}table{{width:100%;border-collapse:collapse;font-size:13px}}th,td{{padding:10px 8px;border-bottom:1px solid var(--line);text-align:right}}th:first-child,td:first-child{{text-align:left}}th{{color:var(--muted)}}.note{{border-left:4px solid var(--amber);background:rgba(251,191,36,.08);color:#fde68a;border-radius:14px;padding:14px 16px;line-height:1.45;font-size:14px}}.src{{font-size:12px;color:var(--muted);word-break:break-all;line-height:1.55}}@media(max-width:1050px){{.span3,.span4,.span5,.span7{{grid-column:span 12}}.dept-grid{{grid-template-columns:1fr}}.hero{{display:block}}}}


.next-steps{{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:10px;margin-top:14px}}.next-steps div{{background:#f8fbff;border:1px solid #e5edf5;border-radius:7px;padding:12px;color:#64748d;font-size:13px;line-height:1.3}}.next-steps b{{display:block;color:#061b31;font-weight:500;margin-bottom:4px}}.decision-status{{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:10px;margin-top:12px}}.decision-status div{{border:1px solid #e5edf5;background:#fff;border-radius:7px;padding:11px}}.decision-status b{{display:block;color:#061b31;font-weight:500;margin-bottom:3px}}.decision-status span{{display:block;color:#64748d;font-size:13px;line-height:1.28}}.guide-grid{{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;margin-top:12px}}.guide-grid div{{background:#f8fbff;border:1px solid #e5edf5;border-radius:6px;padding:12px}}.guide-grid b{{display:block;color:#061b31;font-weight:500;margin-bottom:4px}}.guide-grid span{{display:block;color:#64748d;font-size:13px;line-height:1.3}}.page-guide details,.source-drawer{{margin-top:12px;color:#64748d}}.page-guide summary,.source-drawer summary{{cursor:pointer;color:#533afd;font-weight:500}}@media(max-width:1000px){{.next-steps,.decision-status{{grid-template-columns:repeat(2,minmax(0,1fr))}}}}@media(max-width:760px){{.guide-grid{{grid-template-columns:repeat(2,minmax(0,1fr))}}}}@media(max-width:620px){{.guide-grid,.next-steps,.decision-status{{grid-template-columns:1fr}}}}.dept-grid-card{{padding:20px}}.dept-grid-card h2{{margin:4px 0 6px;font-size:24px;font-weight:500;color:#061b31;letter-spacing:-.02em}}.dept-grid-card .dept-grid{{margin-top:16px}}
.useful-card{{min-height:238px}}.stale-badge{{display:inline-flex;align-items:center;background:#fff7ed;color:#9b6829;border:1px solid #fed7aa;border-radius:4px;padding:4px 7px;font-size:11px;font-weight:600;letter-spacing:.02em;text-transform:uppercase}}.decision-line{{color:#64748d;font-size:14px;line-height:1.35;margin:9px 0 18px;min-height:42px}}.budget-scale{{position:relative;height:10px;background:#e6edf7;border-radius:5px;margin:10px 0 14px;overflow:hidden}}.budget-scale i{{display:block;height:100%;background:linear-gradient(90deg,#533afd,#f96bee);border-radius:5px}}.budget-scale span{{position:absolute;right:0;top:-18px;color:#64748d;font-size:11px;background:#fff;padding:0 3px}}.useful-nums{{grid-template-columns:1fr 1fr 1fr}}.useful-nums div{{min-height:84px}}.useful-nums small{{font-size:11px;line-height:1.25}}.card-action{{margin-top:12px;border-top:1px solid #e5edf5;padding-top:10px;color:#64748d;font-size:12px;line-height:1.35}}@media(max-width:1000px){{.useful-nums{{grid-template-columns:1fr}}}}
.budget-compare{{padding:20px 22px 18px;overflow:hidden}}.chart-head{{display:flex;justify-content:space-between;gap:18px;align-items:flex-start;margin-bottom:14px}}.chart-head h2{{margin:4px 0 4px;font-size:24px;font-weight:500;letter-spacing:-.02em;color:#061b31}}.source-chips{{display:flex;gap:8px;flex-wrap:wrap;justify-content:flex-end}}.source-chips span{{display:inline-flex;align-items:center;border:1px solid #e5edf5;background:#f8fbff;color:#64748d;border-radius:6px;padding:6px 8px;font-size:12px;white-space:nowrap}}.table-scroll{{width:100%;overflow-x:auto;border:1px solid #e5edf5;border-radius:8px}}.compare-table{{width:100%;min-width:880px;font-size:14px;border-collapse:separate;border-spacing:0}}.compare-table th{{font-size:11px;color:#64748d;letter-spacing:.08em;text-transform:uppercase;padding:11px 14px;background:#f8fbff;position:sticky;top:0;z-index:1}}.compare-row{{cursor:pointer;transition:background .12s ease}}.compare-row:hover{{background:#f6f9fc}}.compare-row td{{padding:12px 14px;border-bottom:1px solid #e5edf5;vertical-align:middle}}.dept-name{{width:31%;min-width:245px}}.dept-name b{{display:block;color:#061b31;font-size:15px;font-weight:650;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}.num{{font-variant-numeric:tabular-nums;text-align:right;color:#273951;font-weight:600;white-space:nowrap}}.num span{{display:block;color:#64748d;font-size:11px;margin-top:2px}}.num.actual{{color:#9b6829}}.progress-cell{{min-width:190px;display:grid;grid-template-columns:1fr 46px;gap:10px;align-items:center}}.progress-cell strong{{font-variant-numeric:tabular-nums;color:#273951;text-align:right;font-weight:600}}.progress-line{{height:9px;border-radius:5px;background:#e6edf7;overflow:hidden}}.progress-line i{{height:100%;display:block;border-radius:5px;background:#15be53}}.progress-line i.tight{{background:#9b6829}}.progress-line i.over{{background:#ea2261}}.compare-row.over .dept-name b{{color:#9f1239}}.compare-row.tight .dept-name b{{color:#7c4a03}}.row-action{{width:96px;text-align:right;color:#533afd;font-size:12px;font-weight:500;white-space:nowrap}}.compact-section{{padding:0!important;overflow:hidden}}.compact-section summary{{cursor:pointer;list-style:none;display:flex;justify-content:space-between;gap:16px;align-items:center;padding:18px 20px;color:#061b31}}.compact-section summary::-webkit-details-marker{{display:none}}.compact-section summary b{{font-size:20px;font-weight:500;letter-spacing:-.02em}}.compact-section summary span{{color:#64748d;font-size:13px;line-height:1.3;text-align:right;max-width:520px}}.compact-body{{border-top:1px solid #e5edf5;padding:16px 20px 20px}}.mini-cards{{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:10px;margin-bottom:12px}}.mini-cards div{{border:1px solid #e5edf5;background:#f8fbff;border-radius:6px;padding:10px;color:#64748d;font-size:13px}}.mini-cards b{{display:block;color:#061b31;font-weight:500;margin-bottom:3px}}@media(max-width:1100px){{.chart-head{{display:block}}.legend{{justify-content:flex-start;margin-top:10px}}.compare-table{{min-width:760px}}.budget-compare{{overflow:auto}}.mini-cards{{grid-template-columns:1fr}}.compact-section summary{{display:block}}.compact-section summary span{{display:block;text-align:left;margin-top:4px}}}}
</style><style>.dept{{cursor:pointer;transition:transform .15s ease,border-color .15s ease}}.dept:hover{{transform:translateY(-2px);border-color:#b9b9f9}}.modal{{position:fixed;inset:0;background:rgba(6,27,49,.38);backdrop-filter:blur(5px);display:none;align-items:center;justify-content:center;z-index:50;padding:24px}}.modal.open{{display:flex}}.modal-card{{width:min(1120px,96vw);max-height:90vh;overflow:auto;background:#fff!important;border:1px solid #e5edf5!important;border-radius:10px!important;padding:22px;box-shadow:rgba(50,50,93,.25) 0 30px 45px -30px,rgba(0,0,0,.1) 0 18px 36px -18px!important;color:#273951!important}}.modal-top{{display:flex;justify-content:space-between;gap:16px;align-items:start}}.close{{background:#fff!important;border:1px solid #e5edf5!important;color:#533afd!important;border-radius:6px!important;padding:8px 12px;cursor:pointer}}.proj{{width:120px;background:#fff!important;color:#061b31!important;border:1px solid #e5edf5!important;border-radius:6px!important;padding:8px;text-align:right}}.btn{{background:#533afd!important;border:1px solid #533afd!important;color:#fff!important;border-radius:4px!important;padding:9px 13px;font-weight:400;cursor:pointer}}.mini{{font-size:12px;color:#64748d!important}}</style></head><body><div class='wrap'>
<header class='hero compact-hero'><div><h1>Department budgets</h1><div class='sub'>Budget · actual · remaining. Click a row for evidence.</div></div><div class='header-meta'><span class='pill'>Budget: FY2026 approved</span><span class='pill'>Actuals: {esc(actual_period)}</span><span class='pill'>{esc(source_value_text)}</span></div></header>
<section class='grid'>
<div class='card span12 budget-compare'><div class='chart-head'><div><div class='label'>Budget control</div><h2>Department budget vs actuals</h2><div class='mini'>One row per department. Click any row for evidence lines.</div></div><div class='source-chips'><span>Budget · FY2026 approved</span><span>Actuals · {esc(actual_period)}</span><span>{esc(source_value_text)}</span></div></div><div class='table-scroll'><table class='compare-table'><thead><tr><th>Department</th><th>Budget</th><th>Actual</th><th>Remaining / over</th><th>% used</th><th></th></tr></thead><tbody>{compare_rows}</tbody></table></div></div>
<div class='card span12 compact-section'><details><summary><b>{esc(next_title)}</b><span>{esc(next_not_current_label)} · {esc(next_action_text)}</span></summary><div class='compact-body'><div class='decision-status'><div><b>Use now</b><span>Approved FY2026 department authority.</span></div><div><b>{esc(next_not_current_label)}</b><span>{esc(next_not_current_text)}</span></div><div><b>Next</b><span>{esc(next_action_text)}</span></div></div><div class='sub'>{esc(next_intro)}</div></div></details></div>
<details class='card span12 compact-section'><summary><b>Source and operating summary</b><span>MYOB actuals · FY2026 approved budget · May operating summary</span></summary><div class='compact-body'><div class='mini-cards'><div><b>Income — {esc(summary_period)}</b>{money(data['summary']['income'])}</div><div><b>Spend — {esc(summary_period)}</b>{money(data['summary']['spend'])}</div><div><b>Actual source</b>{esc(source_status_text)}</div><div><b>June pace</b>{esc(pace_note)}</div></div><div class='sub'>{esc(period_note)} Department pages answer ministry/function budget authority. People pages answer named staff cost. Do not add them together without reconciling source basis.</div></div></details>
<div class='card span12'><input class='search' id='search' placeholder='Search department e.g. evangelism, faith, president, youth, ministerial, camp...' oninput='filterCards()'></div>
<div class='card span12 dept-grid-card'><div class='label'>Department decision cards</div><h2>Budget / Spend / Remaining</h2><div class='sub'>The coloured line is actual spend as a percentage of annual budget. If MYOB live extraction is active, this is transaction-backed; otherwise it is legacy indexed detail and must be refreshed before decisions.</div><div class='dept-grid' id='deptGrid'>{cards}</div></div>
<details class='card span12 compact-section'><summary><b>Field budget map + vacancy / savings lens</b><span>Open when reviewing FIELD wage lines or drafting FY2027 projections.</span></summary><div class='compact-body'><div class='mini-cards'><div><b>Answer</b>Use FIELD budget authority and person-cost pages together, not as one number.</div><div><b>Risk</b>Underspend may be vacancy savings or missed ministry capacity.</div><div><b>Source</b>Approved FY2026 FIELD budget + {esc(actual_period)} spend detail.</div></div><table><thead><tr><th>Field line</th><th>FY2026 budget</th><th>Spent — {esc(actual_period)}</th><th>Expected at {esc(elapsed_text)}</th><th>Under / over pace</th><th>Draft FY2027 projection</th></tr></thead><tbody>{field_rows}</tbody></table><br><button class='btn' onclick='saveProjections()'>Save FY2027 projections locally</button> <button class='btn' onclick='exportProjections()'>Download FY2027 projection JSON</button><div class='mini'>Stored in this browser only via localStorage. Export JSON if you want me to turn it into a proper forecast file later.</div></div></details>
<details class='card span12 compact-section'><summary><b>Previous-year history status</b><span>History is intentionally hidden until older workbook/SUN sources are indexed.</span></summary><div class='compact-body'><div class='mini-cards'><div><b>Available now</b>FY2026 budget/spend and {esc(actual_period)} detail.</div><div><b>Not safe yet</b>No 10-year trend from filenames or mixed workbooks.</div><div><b>Next</b>Index 2025–2016 budget/actuals with source rows.</div></div><table><thead><tr><th>Department</th><th>History available now</th><th>Next practical action</th></tr></thead><tbody><tr><td>EVANGELISM</td><td>2026 budget/spend available in this dashboard</td><td>Index 2025–2016 budget/actual workbooks, then show annual budget, actual spend, variance, and over/under flag.</td></tr><tr><td>FIELD / ADMIN / YOUTH / AAV</td><td>2026 budget/spend available in this dashboard</td><td>Use the same department-history schema once historical sources are parsed.</td></tr></tbody></table></div></details>
<details class='card span7 compact-section'><summary><b>All department/function budgets</b><span>Full table for export/checking; main decision view is the card/table above.</span></summary><div class='compact-body'><div class='mini'>{esc(budget_period)} vs actual spend from {esc(actual_period)}. Pace compares spend to the share of the year elapsed, not to a monthly budget.</div><table id='deptTable'><thead><tr><th>Department/function</th><th>Budget</th><th>Spent</th><th>Expected by elapsed year</th><th>Under / over pace</th><th>Used</th><th>Elapsed</th></tr></thead><tbody>{rows}</tbody></table></div></details>
<div class='card span5'><div class='label'>Cash balances visible — {esc(summary_period)}</div><table><thead><tr><th>Account</th><th>Type</th><th>{esc(summary_period)}</th></tr></thead><tbody>{cash}</tbody></table><br><div class='note'>Source: operating summary, {esc(summary_period)}. This is a dashboard for capacity and timing questions. It does not replace approval authority, restricted funding rules, or strategic priority judgement.</div></div>
<details class='card span12 compact-section'><summary><b>Sources and period context</b><span>Open for workbook/PDF path, modification time and generated JSON.</span></summary><div class='compact-body'><div class='src'>Budget period: {esc(budget_period)} · source modified {esc(period.get('budget_source_modified'))}<br>Actual spend period: {esc(actual_period)} · source modified {esc(period.get('actual_source_modified'))}<br>Summary period: {esc(summary_period)} · source modified {esc(period.get('summary_source_modified'))}<br>Detailed department report: {esc(data['source'])}<br>JSON: {esc(str(OUT_JSON))}</div></div></details>
</section></div>
<div class='modal' id='deptModal' onclick='if(event.target.id==="deptModal") closeDept()'><div class='modal-card'><div class='modal-top'><div><div class='label'>Department drilldown</div><h2 id='modalTitle'></h2><div class='sub' id='modalSummary'></div></div><button class='close' onclick='closeDept()'>Close</button></div><br><div class='note'>{esc(drilldown_note)}</div><br><table><thead><tr><th>Line / activity</th><th>Budget</th><th>Spent</th><th>Remaining</th></tr></thead><tbody id='modalRows'></tbody></table></div></div>
<script>
const D={chart};
function fmt(x){{return '$'+Math.round(Math.abs(x)).toLocaleString();}}
function shortName(name){{return name.length>24?name.slice(0,23)+'…':name;}}
function drawUsed(){{
  const c=document.getElementById('used'); if(!c) return; const ctx=c.getContext('2d');
  const rows=D.departments.slice(0,14); c.width=c.clientWidth*2; c.height=400*2; ctx.scale(2,2);
  const W=c.clientWidth,p=190,rowH=25,max=140;
  ctx.clearRect(0,0,W,400); ctx.font='12px Source Sans 3, system-ui'; ctx.textBaseline='middle';
  rows.forEach((d,i)=>{{
    const y=42+i*rowH,w=Math.min(max,d.used)/max*(W-p-56);
    ctx.fillStyle='#273951'; ctx.textAlign='right'; ctx.fillText(shortName(d.name),p-12,y+6);
    ctx.fillStyle='#e6edf7'; ctx.fillRect(p,y,W-p-56,12);
    ctx.fillStyle=d.used>100?'#ea2261':d.used>85?'#9b6829':'#15be53'; ctx.fillRect(p,y,w,12);
    ctx.fillStyle='#64748d'; ctx.textAlign='left'; ctx.fillText(Math.round(d.used)+'%',p+w+7,y+6);
  }});
}}
function filterCards(){{const q=document.getElementById('search').value.toLowerCase();document.querySelectorAll('.dept').forEach(el=>{{el.style.display=el.dataset.name.includes(q)?'block':'none'}})}}
function openDept(name){{const d=D.allDepartments.find(x=>x.name===name); if(!d) return; document.getElementById('modalTitle').textContent=d.name; const rem=d.remaining<0?'($'+Math.round(Math.abs(d.remaining)).toLocaleString()+')':'$'+Math.round(d.remaining).toLocaleString(); document.getElementById('modalSummary').textContent='Budget ({budget_period}) $'+Math.round(d.budget).toLocaleString()+' • Spent ({actual_period}) $'+Math.round(d.spent).toLocaleString()+' • Remaining vs FY2026 budget '+rem+' • '+Math.round(d.used)+'% used'; const rows=(d.lines||[]).map(x=>'<tr><td>'+x.line+'</td><td>'+fmt(x.budget)+'</td><td>'+fmt(x.spent)+'</td><td class="'+(x.remaining<0?'bad':'good')+'">'+(x.remaining<0?'('+fmt(x.remaining)+')':fmt(x.remaining))+'</td></tr>').join('') || '<tr><td colspan="4">No line-level detail available in the selected Velixo report for {actual_period}.</td></tr>'; document.getElementById('modalRows').innerHTML=rows; document.getElementById('deptModal').classList.add('open')}}
function closeDept(){{document.getElementById('deptModal').classList.remove('open')}}
function saveProjections(){{const obj={{savedAt:new Date().toISOString(), field:{{}}}}; document.querySelectorAll('.proj').forEach(i=>obj.field[i.dataset.fieldLine]=Number(i.value||0)); localStorage.setItem('snswFieldProjections', JSON.stringify(obj)); alert('Field projections saved locally in this browser.')}}
function loadProjections(){{try{{const obj=JSON.parse(localStorage.getItem('snswFieldProjections')||'{{}}'); document.querySelectorAll('.proj').forEach(i=>{{if(obj.field && obj.field[i.dataset.fieldLine]!==undefined) i.value=obj.field[i.dataset.fieldLine];}})}}catch(e){{}}}}
function exportProjections(){{const obj={{savedAt:new Date().toISOString(), field:{{}}}}; document.querySelectorAll('.proj').forEach(i=>obj.field[i.dataset.fieldLine]=Number(i.value||0)); const blob=new Blob([JSON.stringify(obj,null,2)],{{type:'application/json'}}); const a=document.createElement('a'); a.href=URL.createObjectURL(blob); a.download='field-budget-projections.json'; a.click();}}
drawUsed();loadProjections();
</script></body></html>"""

def budget_departments_from_2026_budget():
    # Approved control totals from Final budget 2026.pdf, presented to the Board on 15 Feb 2026.
    # Do not use Budgets 2026.xlsx here; it is an older MYOB/import basis and does not tie to the approved PDF.
    depts=[]
    for name, amount in APPROVED_DEPARTMENT_BUDGETS.items():
        lines=[{'line':line,'budget':budget,'spent':0.0,'remaining':budget} for line,budget in APPROVED_DEPARTMENT_LINES.get(name, [(name, amount)])]
        depts.append({'name':name,'budget':amount,'spent':0.0,'remaining':amount,'used_pct':None,'status':'ok','income_budget':0.0,'income_actual':0.0,'lines':lines})
    return depts

def adventist_community_services_from_income(path):
    """Create ACS budget from incoming donated/community-services income lines.

    This is intentionally synthetic: the 2026 budget workbook has no ACS department,
    so Kyle wants ACS budget capacity to equal donated income coded to ACS/community
    service accounts in the selected operating report.
    """
    terms=['adventist community service','adventist community services','snsw community services','snsw adventist community services','community service merchandise']
    wb=load_workbook(path,read_only=True,data_only=True,keep_links=False)
    if 'Rpt B-Functions ' not in wb.sheetnames:
        return None
    ws=wb['Rpt B-Functions ']
    income=0.0; spent=0.0; lines=[]
    for row in ws.iter_rows(min_row=11, values_only=True):
        desc=str(row[3] or '').strip()
        if not desc:
            continue
        text=desc.lower()
        if not any(t in text for t in terms):
            continue
        budget=n(row[4] if len(row)>4 else 0)
        actual=n(row[5] if len(row)>5 else 0)
        is_income=('income' in text or 'sale of goods' in text or actual>0)
        if is_income:
            income += max(0.0, actual)
            lines.append({'line':desc + ' (incoming allocation)', 'budget':max(0.0, actual), 'spent':0.0, 'remaining':max(0.0, actual)})
        else:
            spent += abs(actual)
            lines.append({'line':desc, 'budget':0.0, 'spent':abs(actual), 'remaining':-abs(actual)})
    if income==0 and spent==0:
        return {'name':'ADVENTIST COMMUNITY SERVICES','budget':0.0,'spent':0.0,'remaining':0.0,'used_pct':None,'status':'ok','income_budget':0.0,'income_actual':0.0,'lines':[{'line':'No ACS/community-services donations found in the selected Velixo operating report for the stated actual period','budget':0.0,'spent':0.0,'remaining':0.0}]}
    remaining=income-spent
    used=(spent/income*100) if income else None
    return {'name':'ADVENTIST COMMUNITY SERVICES','budget':income,'spent':spent,'remaining':remaining,'used_pct':used,'status':'over' if remaining<0 else 'tight' if used and used>85 else 'ok','income_budget':income,'income_actual':income,'lines':lines[:12]}

def load_myob_report_if_current():
    if not MYOB_REPORT.exists():
        return None
    try:
        data=json.loads(MYOB_REPORT.read_text())
    except Exception:
        return None
    source_kind=(data.get('period_context') or {}).get('source_kind')
    if source_kind != 'myob_live_gl_cache':
        return None
    departments=data.get('departments') or []
    if not departments:
        return None
    data['summary']=parse_summary()
    return data

def main():
    OUT_DIR.mkdir(parents=True,exist_ok=True)
    myob_data=load_myob_report_if_current()
    if myob_data:
        myob_data['generated_at'] = datetime.now().isoformat(timespec='seconds')
        ensure_theme_file(OUT_DIR)
        OUT_JSON.write_text(json.dumps(myob_data,indent=2),encoding='utf-8')
        OUT_HTML.write_text(apply_stripe_theme(render(myob_data)),encoding='utf-8')
        print(OUT_HTML)
        return
    source=pick_source()
    actual_depts=parse_functions(source)
    budget_depts=budget_departments_from_2026_budget()
    acs=adventist_community_services_from_income(source)
    if acs:
        budget_depts=[d for d in budget_depts if d['name']!='ADVENTIST COMMUNITY SERVICES'] + [acs]
    actual_by_name={d['name']:d for d in actual_depts}
    for d in budget_depts:
        a=actual_by_name.get(d['name'])
        if a:
            d['spent']=a['spent']
            d['remaining']=d['budget']-d['spent']
            d['used_pct']=(d['spent']/d['budget']*100) if d['budget'] else None
            d['status']='over' if d['remaining']<0 else 'tight' if d['used_pct'] and d['used_pct']>85 else 'ok'
            # If actual report has useful line detail, merge it into drilldown after budget lines.
            if a.get('lines'):
                d['lines']=(d['lines'] + a['lines'])[:12]
    depts=sorted(budget_depts, key=lambda x:x['budget'], reverse=True)
    summary=parse_summary()
    data={'generated_at':datetime.now().isoformat(timespec='seconds'),'source':str(APPROVED_BUDGET_PDF)+' + approval email note: '+str(APPROVED_BUDGET_EMAIL_NOTE)+' + actual detail source: '+str(source),'source_modified':datetime.fromtimestamp(APPROVED_BUDGET_PDF.stat().st_mtime).isoformat(timespec='minutes'),'period_context':build_period_context(source),'departments':depts,'summary':summary}
    OUT_JSON.write_text(json.dumps(data,indent=2),encoding='utf-8')
    ensure_theme_file(OUT_DIR)
    OUT_HTML.write_text(apply_stripe_theme(render(data)),encoding='utf-8')
    print(OUT_HTML)
if __name__=='__main__': main()
