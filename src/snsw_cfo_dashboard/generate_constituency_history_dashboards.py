#!/usr/bin/env python3
"""Generate starter constituency history dashboards from the local session catalogue.

Read-only against the OneDrive Session SNSW folders. This deliberately does not
parse all PDFs deeply; it scans file metadata, names, and shallow workbook sheet
names to build a source/evidence catalogue and starter investigation queue.
"""
from __future__ import annotations

import html
import json
import mimetypes
import os
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - generator still works without xlsx inspection
    load_workbook = None

WORKSPACE = Path("/Users/snswcommunications/Hermes-CFO")
OUT_DIR = WORKSPACE / "briefings" / "dashboards"
SESSION_ROOT = Path("/Users/snswcommunications/Library/CloudStorage/OneDrive-Seventh-dayAdventistChurch(SouthPacific)/Files - SNSW-CFO - CFO/CFO/Session SNSW")
YEARS = ["2005", "2008", "2011", "2014", "2017", "2021", "2025"]
OUT_JSON = OUT_DIR / "constituency-history-data.json"
OUT_MASTER = OUT_DIR / "master-report-constituency-history.html"
OUT_INVESTIGATIONS = OUT_DIR / "constituency-investigations-layer.html"
THEME_CSS = "stripe-cfo-theme.css"

SOURCE_KEYWORDS = {
    "master_report": ["session report", "treasurer", "cfo", "graphs", "charts", "tithe data", "financial report"],
    "finance_statements": ["financial", "income", "balance", "cashflow", "working capital", "audit", "session report"],
    "investigation": ["draft", "old", "updated", "copy", "version", "final", "opn", "audit opinion", "cashflow", "working capital"],
    "constituency_story": ["member", "delegate", "constituency", "session", "tithe", "offering", "financial", "treasurer", "cfo"],
}

CLAIMS = [
    {
        "id": "membership-delegate-spine",
        "question": "What are the published membership and delegate counts at each session?",
        "status": "needs extraction",
        "why": "This is the master report spine, but the current pass only catalogues source files. Counts should be lifted from published report pages/slides, not inferred.",
        "search_terms": ["member", "membership", "delegate", "constituency", "session report"],
        "priority": "high",
    },
    {
        "id": "tithe-growth-claims",
        "question": "Do tithe-growth claims reconcile across the CFO slides, graph workbooks, and published reports?",
        "status": "queued",
        "why": "Several Tithe Data / Graphs & Charts workbooks exist across 2008-2025 and should be compared before using growth language.",
        "search_terms": ["tithe", "graphs", "charts", "CFO", "treasurer"],
        "priority": "high",
    },
    {
        "id": "draft-vs-final-risk",
        "question": "Which finance statements are final/published versus drafts, old copies, or updated sheets?",
        "status": "catalogued",
        "why": "The 2021 and 2025 lanes include final-looking reports plus OLD/draft/update artefacts; evidence labels need clear source status.",
        "search_terms": ["draft", "old", "final", "updated", "audit opinion"],
        "priority": "medium",
    },
    {
        "id": "entity-boundary",
        "question": "Are SNC, SNE, SNU, school, company and conference figures being mixed?",
        "status": "queued",
        "why": "The folders contain multiple reporting entities. Challenge claims that use a single figure without entity boundary.",
        "search_terms": ["SNC", "SNE", "SNU", "school", "conference", "trust"],
        "priority": "high",
    },
    {
        "id": "working-capital-cashflow",
        "question": "Where working-capital or cashflow claims appear, do they point to the matching statement and year?",
        "status": "queued",
        "why": "Working capital and cashflow files appear explicitly in 2021/2025; they should be evidence-backed before narrative use.",
        "search_terms": ["working capital", "cashflow", "cash flow", "balance sheet"],
        "priority": "medium",
    },
]


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def file_url(path: Path) -> str:
    return "file://" + quote(str(path))


def score_file(rel: str, name: str) -> dict[str, int]:
    text = f"{rel} {name}".lower()
    return {k: sum(1 for term in terms if term in text) for k, terms in SOURCE_KEYWORDS.items()}


def classify_source(path: Path) -> str:
    text = str(path).lower()
    if any(t in text for t in ["draft", "old", "copy", "version", "updated"]):
        return "working / draft / update"
    if any(t in text for t in ["final", "audit opinion", "audited"]):
        return "published / final-looking"
    if any(t in text for t in ["session report", "treasurer", "cfo", "graphs", "charts", "tithe data", "financial report"]):
        return "candidate evidence"
    return "catalogue item"


def inspect_xlsx(path: Path) -> dict:
    if load_workbook is None or path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return {}
    try:
        wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        previews = []
        for sh in wb.sheetnames[:6]:
            ws = wb[sh]
            sample = []
            for row in ws.iter_rows(min_row=1, max_row=6, max_col=6, values_only=True):
                vals = [str(v).strip() for v in row if v not in (None, "")]
                if vals:
                    sample.append(" | ".join(vals[:5]))
                if len(sample) >= 3:
                    break
            previews.append({"sheet": sh, "sample": sample})
        return {"sheets": wb.sheetnames, "preview": previews}
    except Exception as exc:
        return {"xlsx_warning": str(exc)}


def catalogue() -> dict:
    generated_at = now_iso()
    years = []
    all_evidence = []
    warnings = []
    if not SESSION_ROOT.exists():
        warnings.append(f"Session root not found: {SESSION_ROOT}")

    for year in YEARS:
        ydir = SESSION_ROOT / year
        files = [p for p in ydir.rglob("*") if p.is_file()] if ydir.exists() else []
        suffix_counts = Counter((p.suffix.lower() or "[no extension]") for p in files)
        file_items = []
        ranked = []
        for p in files:
            rel = str(p.relative_to(SESSION_ROOT))
            st = p.stat()
            scores = score_file(rel, p.name)
            item = {
                "id": re.sub(r"[^a-zA-Z0-9]+", "-", rel).strip("-").lower(),
                "year": year,
                "name": p.name,
                "relative_path": rel,
                "path": str(p),
                "url": file_url(p),
                "extension": p.suffix.lower() or "[no extension]",
                "mime_guess": mimetypes.guess_type(str(p))[0],
                "size_bytes": st.st_size,
                "modified": datetime.fromtimestamp(st.st_mtime).isoformat(timespec="minutes"),
                "source_status": classify_source(p),
                "scores": scores,
            }
            if p.suffix.lower() in {".xlsx", ".xlsm"} and max(scores.values() or [0]) > 0:
                item["workbook_inspection"] = inspect_xlsx(p)
            file_items.append(item)
            rank = scores["master_report"] * 3 + scores["finance_statements"] * 2 + scores["investigation"]
            if rank:
                ranked.append((rank, item))
                all_evidence.append(item)
        ranked.sort(key=lambda x: (-x[0], x[1]["relative_path"].lower()))
        years.append({
            "year": year,
            "folder": str(ydir),
            "folder_url": file_url(ydir),
            "exists": ydir.exists(),
            "file_count": len(files),
            "extension_counts": dict(sorted(suffix_counts.items())),
            "top_evidence": [item for _, item in ranked[:12]],
            "all_files": sorted(file_items, key=lambda x: x["relative_path"].lower()),
            "trend_point": {
                "session_year": year,
                "published_member_count": None,
                "published_delegate_count": None,
                "source_state": "catalogued; counts not extracted in this starter pass",
                "primary_evidence_ids": [item["id"] for _, item in ranked[:4]],
            },
        })

    claim_evidence = []
    for claim in CLAIMS:
        terms = [t.lower() for t in claim["search_terms"]]
        matches = []
        for ev in all_evidence:
            hay = f"{ev['relative_path']} {ev['name']} {ev['source_status']}".lower()
            hit = sum(1 for term in terms if term in hay)
            if hit:
                matches.append((hit, ev))
        matches.sort(key=lambda x: (-x[0], x[1]["year"], x[1]["relative_path"].lower()))
        claim_evidence.append({**claim, "evidence_ids": [ev["id"] for _, ev in matches[:10]], "evidence": [ev for _, ev in matches[:10]]})

    return {
        "generated_at": generated_at,
        "generator": str(WORKSPACE / "tools" / "dashboard" / "generate_constituency_history_dashboards.py"),
        "source_root": str(SESSION_ROOT),
        "source_root_url": file_url(SESSION_ROOT),
        "method": "Folder catalogue + filename/status heuristics + shallow workbook sheet previews for candidate xlsx/xlsm files. No PDF deep parsing and no source mutation.",
        "years": years,
        "claims": claim_evidence,
        "warnings": warnings,
        "all_evidence_count": len(all_evidence),
    }


def h(text: object) -> str:
    return html.escape("" if text is None else str(text), quote=True)


def pill(label: str, cls: str = "") -> str:
    return f'<span class="pill {cls}">{h(label)}</span>'


def evidence_link(item: dict, label: str | None = None) -> str:
    return f'<a class="src-link" href="{h(item["url"])}" title="{h(item["path"])}">{h(label or item["name"])}</a>'


def base_css() -> str:
    return """
    .hero{display:grid;grid-template-columns:1.35fr .65fr;gap:18px;align-items:stretch;margin:22px 0}.hero .card{min-height:220px}.eyebrow{color:#533afd;font-weight:500;letter-spacing:.11em;text-transform:uppercase;font-size:12px}.deck{font-size:18px;line-height:1.55;color:#42526a;max-width:930px}.local-tabs{display:flex;gap:8px;flex-wrap:wrap;margin-top:18px}.local-tabs a{border:1px solid var(--stripe-line);background:#fff;border-radius:999px;padding:8px 12px;box-shadow:rgba(23,23,23,.04) 0 2px 8px}.timeline{display:grid;grid-template-columns:repeat(7,minmax(120px,1fr));gap:10px;margin-top:18px}.year-card{padding:14px;border:1px solid var(--stripe-line);border-radius:8px;background:#fff;box-shadow:rgba(23,23,23,.04) 0 2px 8px}.year-card strong{display:block;color:#061b31;font-size:22px;font-weight:400}.year-card small{display:block;color:#64748d;margin-top:4px}.source-list{display:grid;gap:8px;margin-top:12px}.source-row{display:grid;grid-template-columns:90px 1fr auto;gap:10px;align-items:center;padding:10px 0;border-bottom:1px solid var(--stripe-line)}.source-row:last-child{border-bottom:0}.src-link{font-family:ui-monospace,SFMono-Regular,Menlo,monospace;font-size:12px;word-break:break-word}.table-wrap{overflow:auto;border:1px solid var(--stripe-line);border-radius:8px;background:#fff}.trend-table td,.trend-table th{padding:11px 12px;text-align:left}.muted{color:#64748d}.note-box{padding:14px;border:1px solid #d6d9fc;background:linear-gradient(90deg,#f0efff,#fff);border-radius:8px}.question-ui{display:grid;grid-template-columns:1fr auto;gap:10px;margin:14px 0}.question-ui input{padding:13px 14px;font-size:15px}.answer{min-height:82px}.evidence-chip{display:inline-flex;margin:3px 5px 3px 0;padding:5px 7px;border:1px solid var(--stripe-line);border-radius:6px;background:#f8fbff;font-size:12px}.queue{display:grid;gap:12px}.queue-item{padding:14px;border:1px solid var(--stripe-line);border-radius:8px;background:#fff}.queue-head{display:flex;justify-content:space-between;gap:12px;align-items:flex-start}.status{font-size:11px;text-transform:uppercase;letter-spacing:.08em;color:#64748d}.mini-chart{height:180px;display:flex;align-items:end;gap:10px;padding:12px;border:1px solid var(--stripe-line);border-radius:8px;background:linear-gradient(180deg,#fff,#f8fbff)}.bar-col{flex:1;display:flex;flex-direction:column;align-items:center;gap:7px}.bar-col i{display:block;width:100%;min-height:10px;border-radius:6px 6px 2px 2px;background:linear-gradient(180deg,#7c6bff,#533afd)}.bar-col span{font-size:11px;color:#64748d}.footnote{margin-top:22px}.two-col{display:grid;grid-template-columns:1fr 1fr;gap:18px}@media(max-width:1000px){.hero,.two-col{grid-template-columns:1fr}.timeline{grid-template-columns:repeat(2,minmax(0,1fr))}.source-row{grid-template-columns:1fr}.question-ui{grid-template-columns:1fr}}
    """


def render_master(data: dict) -> str:
    total_files = sum(y["file_count"] for y in data["years"])
    trend_rows = "".join(
        f"<tr><td><b>{h(y['year'])}</b></td><td class='muted'>Not extracted</td><td class='muted'>Not extracted</td><td>{h(y['file_count'])}</td><td>{', '.join(h(k)+': '+h(v) for k,v in y['extension_counts'].items())}</td><td>{' '.join(evidence_link(ev, ev['name']) for ev in y['top_evidence'][:3])}</td></tr>"
        for y in data["years"]
    )
    timeline = "".join(
        f"<div class='year-card'><strong>{h(y['year'])}</strong><small>{h(y['file_count'])} catalogue files</small><small>{h(len(y['top_evidence']))} candidate evidence items</small><a href='{h(y['folder_url'])}'>Open folder</a></div>"
        for y in data["years"]
    )
    source_row_parts = []
    for y in data["years"]:
        year_sources = "".join(
            "<div>" + evidence_link(ev) + " <span class='muted'>— " + h(ev["source_status"]) + "</span></div>"
            for ev in y["top_evidence"][:5]
        )
        source_row_parts.append(
            f"<div class='source-row'><b>{h(y['year'])}</b><div>{year_sources}</div><a href='{h(y['folder_url'])}'>Folder</a></div>"
        )
    source_rows = "".join(source_row_parts)
    bars = "".join(f"<div class='bar-col'><i style='height:{max(12,min(160,y['file_count']*2))}px'></i><span>{h(y['year'])}</span></div>" for y in data["years"])
    return f"""<!doctype html><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Master Report — Constituency Reported History</title><style>{base_css()}</style><link rel="stylesheet" href="{THEME_CSS}"></head><body class="stripe-cfo"><main class="wrap"><header class="hero"><section class="card"><div class="eyebrow">Master report / published-history spine</div><h1>Constituency reported history</h1><p class="deck">Starter master page for the published member/delegate trend spine across Session SNSW folders. This is intentionally not an investigation page: it separates the published trend table from the challenge queue and labels missing extracted values honestly.</p><div class="local-tabs"><a href="master-report-constituency-history.html">Master report</a><a href="constituency-investigations-layer.html">Investigations layer</a><a href="email-intelligence-dashboard.html">Email Intelligence / Source Search</a><a href="constituency-history-data.json">JSON data</a><a href="{h(data['source_root_url'])}">Source root</a></div></section><aside class="card"><div class="label">Generated</div><div class="value">{h(data['generated_at'][:10])}</div><p class="small">{h(data['method'])}</p>{pill('No source mutation','good')} {pill('PDF deep parse skipped','warn')}</aside></header><section class="grid"><div class="card span3"><div class="label">Session folders</div><div class="value">{len(data['years'])}</div><p class="small">2005 → 2025</p></div><div class="card span3"><div class="label">Catalogue files</div><div class="value">{total_files}</div><p class="small">Metadata only, local OneDrive paths</p></div><div class="card span3"><div class="label">Evidence candidates</div><div class="value">{h(data['all_evidence_count'])}</div><p class="small">Filename/status ranked</p></div><div class="card span3"><div class="label">Published counts</div><div class="value warn">Queued</div><p class="small">Membership/delegate values require targeted extraction</p></div></section><section class="card" style="margin-top:18px"><h2>Trend spine status</h2><div class="timeline">{timeline}</div></section><section class="grid" style="margin-top:18px"><div class="card span8"><h2>Published member/delegate trend spine</h2><p class="small">Starter table. Populate count fields only from published report evidence; do not infer from financial files.</p><div class="table-wrap"><table class="trend-table"><thead><tr><th>Session</th><th>Published members</th><th>Published delegates</th><th>Files</th><th>Catalogue mix</th><th>Clickable source affordances</th></tr></thead><tbody>{trend_rows}</tbody></table></div></div><div class="card span4"><h2>Catalogue density</h2><div class="mini-chart">{bars}</div><p class="small">File-count bars are not performance measures; they just show source-density by session folder.</p></div></section><section class="card" style="margin-top:18px"><h2>Primary source lane</h2><div class="source-list">{source_rows}</div></section><section class="note-box footnote"><b>Use rule:</b> This page is the clean master-history layer. Claims that need testing belong in the <a href="constituency-investigations-layer.html">Investigations Layer</a>. Source links are local <code>file://</code> affordances and may require opening from Finder/browser permissions.</section></main></body></html>"""


def render_investigations(data: dict) -> str:
    claim_cards = []
    for c in data["claims"]:
        evlinks = "".join(f"<span class='evidence-chip'>{evidence_link(ev, ev['year']+' / '+ev['name'])}</span>" for ev in c["evidence"][:8]) or "<span class='muted'>No candidate evidence matched.</span>"
        cls = "warn" if c["priority"] == "high" else ""
        claim_cards.append(f"<article class='queue-item' data-question='{h(' '.join(c['search_terms']) + ' ' + c['question'])}'><div class='queue-head'><div><div class='status'>{h(c['status'])}</div><h3>{h(c['question'])}</h3></div>{pill(c['priority'], cls)}</div><p class='small'>{h(c['why'])}</p><div>{evlinks}</div></article>")
    queue = "".join(claim_cards)
    data_script = json.dumps(data["claims"], ensure_ascii=False)
    return f"""<!doctype html><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Investigations Layer — Constituency History</title><style>{base_css()}</style><link rel="stylesheet" href="{THEME_CSS}"></head><body class="stripe-cfo"><main class="wrap"><header class="hero"><section class="card"><div class="eyebrow">Challenge claims / evidence queue</div><h1>Constituency investigations layer</h1><p class="deck">Starter investigation page for testing claims against the local Session SNSW catalogue. ChatGPT-style question box runs locally in-browser over the generated JSON evidence queue; it does not call an API and does not mutate sources.</p><div class="local-tabs"><a href="master-report-constituency-history.html">Master report</a><a href="constituency-investigations-layer.html">Investigations layer</a><a href="email-intelligence-dashboard.html">Email Intelligence / Source Search</a><a href="constituency-history-data.json">JSON data</a><a href="{h(data['source_root_url'])}">Source root</a></div></section><aside class="card"><div class="label">Evidence mode</div><div class="value">Local</div><p class="small">Catalogue-backed prompts: ask about member counts, draft/final risk, tithe claims, entity boundary, cashflow/working capital.</p>{pill('Prototype UI','warn')} {pill('No network','good')}</aside></header><section class="grid"><div class="card span8"><h2>Ask the local evidence queue</h2><div class="question-ui"><input id="q" placeholder="Ask: Which files support delegate counts? Which 2025 reports are draft vs final?" autofocus><button id="ask">Ask</button></div><div class="answer sources" id="answer">Ask a question to filter the generated source catalogue. This is a starter affordance, not an LLM answer.</div></div><div class="card span4"><h2>Investigation rules</h2><ul class="small"><li>Separate published trend values from challenge claims.</li><li>Use report/source labels before narrative.</li><li>Do not mix SNC/SNE/SNU without entity label.</li><li>Prefer final/published-looking evidence over drafts.</li><li>Escalate missing member/delegate counts for targeted extraction.</li></ul></div></section><section class="two-col" style="margin-top:18px"><div class="card"><h2>Claim challenge queue</h2><div class="queue">{queue}</div></div><div class="card"><h2>Evidence extraction backlog</h2><div class="note-box"><b>Next targeted extraction:</b> Open the top Session Report / Treasurer / CFO evidence per year and record published membership + delegate count, page/slide reference, entity label, and exact quote. Avoid bulk PDF parsing until a specific claim requires it.</div><h3>Suggested fields</h3><pre>{{
  "session_year": "2025",
  "claim": "Published membership count",
  "value": null,
  "source_file": "...",
  "page_or_slide": null,
  "quote": null,
  "confidence": "unverified | verified"
}}</pre><h3>Known risk lanes</h3><p class="small">Draft/final ambiguity, entity boundary, old/updated versions, financial-statement versus constituency narrative, and graph workbook reconciliation.</p></div></section></main><script>const CLAIMS={data_script};
function esc(s){{return String(s||'').replace(/[&<>]/g,m=>({{'&':'&amp;','<':'&lt;','>':'&gt;'}}[m]));}}
function ask(){{const q=document.getElementById('q').value.toLowerCase().trim();const out=document.getElementById('answer');if(!q){{out.textContent='Type a question first.';return;}}const terms=q.split(/[^a-z0-9]+/).filter(Boolean);const scored=CLAIMS.map(c=>{{const hay=(c.question+' '+c.why+' '+c.search_terms.join(' ')+' '+(c.evidence||[]).map(e=>e.relative_path+' '+e.source_status).join(' ')).toLowerCase();return [terms.reduce((n,t)=>n+(hay.includes(t)?1:0),0),c];}}).filter(x=>x[0]>0).sort((a,b)=>b[0]-a[0]).slice(0,4);if(!scored.length){{out.innerHTML='No catalogue-backed match. Try terms like <b>member</b>, <b>delegate</b>, <b>draft</b>, <b>SNC</b>, <b>tithe</b>, <b>cashflow</b>.';return;}}out.innerHTML=scored.map(([score,c])=>`<div style="margin-bottom:12px"><b>${{esc(c.question)}}</b><br><span class="muted">Status: ${{esc(c.status)}} · Priority: ${{esc(c.priority)}} · score ${{score}}</span><br>${{esc(c.why)}}<br>${{(c.evidence||[]).slice(0,5).map(e=>`<a class="src-link" href="${{e.url}}">${{esc(e.year+' / '+e.name)}}</a>`).join('<br>')||'<span class="muted">No evidence links</span>'}}</div>`).join('');}}
document.getElementById('ask').addEventListener('click',ask);document.getElementById('q').addEventListener('keydown',e=>{{if(e.key==='Enter')ask();}});</script></body></html>"""


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    data = catalogue()
    OUT_JSON.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    OUT_MASTER.write_text(render_master(data), encoding="utf-8")
    OUT_INVESTIGATIONS.write_text(render_investigations(data), encoding="utf-8")
    print(f"Wrote {OUT_JSON}")
    print(f"Wrote {OUT_MASTER}")
    print(f"Wrote {OUT_INVESTIGATIONS}")


if __name__ == "__main__":
    main()
