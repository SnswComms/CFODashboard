#!/usr/bin/env python3
"""Local pastor allowance email dashboard + safe send trigger.

Serves /Users/snswcommunications/Hermes-CFO/briefings/dashboards and exposes
local API endpoints that pull recipient/FTB data from the Morpheus benefits
tracker, then send via snswfinance@adventist.bot using Morpheus' existing
/api/email/send endpoint.
"""
from __future__ import annotations

import json
import os
import re
import sys
from datetime import date, datetime
from html import escape
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook

ROOT = Path("/Users/snswcommunications")
CFO = ROOT / "Hermes-CFO"
DASH = CFO / "briefings" / "dashboards"
OD = ROOT / "Library/CloudStorage/OneDrive-Seventh-dayAdventistChurch(SouthPacific)"
MORPHEUS = "http://100.87.6.30:8011"  # Morpheus Tailscale IP; Python urllib does not reliably resolve .local
PATHWAYS = OD / "Files - SNSW-Finance - Finance/Payroll/Remuneration Letters/2026/Remuneration/Ministers FBB 2026/Pastors Pathways .xlsx"
BOOK_2025 = OD / "Files - SNSW-Finance - Finance/Payroll/Pay Runs/PR 29-03-2025 Not Printed*/Book and Equipment Ministers.xlsx"
PDE_POLICY = OD / "Files - SNSW-Finance - Finance/Payroll/PR Procedures/Updated DWS Rates 2026/Assistance for Ministers 2026/Professional Development and Equipment Allowance.docx"

CONFIRM_TOKEN = "SEND_PASTOR_ALLOWANCES"
ALLOW_LIVE_SEND = os.environ.get("SNSW_ALLOWANCE_EMAIL_LIVE_SEND", "").strip().lower() in {"1", "true", "yes"}

# Recipients explicitly excluded from the allowance email workflow.
# Keep this in code so refreshes do not reintroduce old staff or allowance-ineligible roles.
EXCLUDED_ALLOWANCE_NAMES = {
    "rangi eiao",        # old staff / moved away
    "lorenzo berry",    # old staff / moved away
    "sharee greenfield",# maternity leave — not currently receiving these allowances
}
EXCLUDED_ALLOWANCE_ASSIGNMENTS = {
    "chaplain",       # school chaplains generally do not receive these minister allowance emails
}
ALLOWANCE_ELIGIBLE_NAME_EXCEPTIONS = {
    "nick chan",      # full pastor serving as chaplain; likely still allowance-eligible
    "nic chan",       # Pathways spelling
}


def _money(v: Any) -> str:
    try:
        n = float(v or 0)
    except Exception:
        n = 0.0
    s = f"${abs(n):,.2f}"
    return f"({s})" if n < 0 else s


def _norm_name(s: str) -> str:
    s = (s or "").lower()
    # Source spelling differences across Pathways, MYOB contact records, and benefits tracker.
    for old, new in {
        "lorangi": "rangi",
        "lawman": "lawnan",
        "olivia": "oliva",
        "issac": "isaac",
        "nic ": "nick ",
        "nicolas": "nick",
        "nicholas": "nick",
    }.items():
        s = s.replace(old, new)
    s = re.sub(r"[^a-z ]+", " ", s)
    parts = [p for p in s.split() if p not in {"pr", "pastor", "elder"}]
    return " ".join(parts)


def _name_key(s: str) -> tuple[str, str]:
    parts = _norm_name(s).split()
    if not parts:
        return ("", "")
    return (parts[0], parts[-1])


def http_json(method: str, url: str, payload: dict | None = None, timeout: int = 40) -> dict:
    data = None if payload is None else json.dumps(payload).encode("utf-8")
    req = Request(url, data=data, method=method, headers={"Content-Type": "application/json"})
    try:
        with urlopen(req, timeout=timeout) as r:
            return json.loads(r.read().decode("utf-8"))
    except HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {e.code} {url}: {body[:500]}") from e
    except URLError as e:
        raise RuntimeError(f"Cannot reach {url}: {e}") from e


def load_pathways() -> list[dict]:
    wb = load_workbook(PATHWAYS, read_only=True, data_only=True, keep_links=False)
    ws = wb["2026"]
    pastors = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = row[1] if len(row) > 1 else None
        if not name or str(name).strip().lower() in {"churches", "name"}:
            continue
        district = row[9] if len(row) > 9 else ""
        assignment = row[10] if len(row) > 10 else ""
        churches = row[11] if len(row) > 11 else ""
        if not district and not assignment and not churches:
            continue
        pastors.append({
            "name": str(name).strip(),
            "district": str(district or "").strip(),
            "assignment": str(assignment or "").strip(),
            "churches": str(churches or "").strip(),
            "ministry_years_total": row[4] if len(row) > 4 else None,
            "snsw_years": row[6] if len(row) > 6 else None,
        })
    return pastors


def load_book_2025() -> dict[tuple[str, str], dict]:
    # Historical/current workbook shape gives per-minister B&E budget/spend/balance.
    import glob
    matches = glob.glob(str(BOOK_2025))
    if not matches:
        return {}
    wb = load_workbook(matches[0], read_only=True, data_only=True, keep_links=False)
    ws = wb["Summary"]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        ref = row[2] if len(row) > 2 else None
        who = row[5] if len(row) > 5 else None
        if str(ref or "").strip().upper() != "TOTAL" or not who:
            continue
        who_s = re.sub(r"\([^)]*\)", "", str(who)).strip()
        if "," in who_s:
            last, first = [x.strip() for x in who_s.split(",", 1)]
            name = f"{first} {last}"
        else:
            name = who_s
        out[_name_key(name)] = {
            "source_name": who_s,
            "book_budget_2025": float(row[6] or 0),
            "first_year_intern_2025": float(row[7] or 0),
            "professional_dev_2025": float(row[8] or 0),
            "book_spent_2025": float(row[9] or 0),
            "book_balance_2025": float(row[10] or 0),
        }
    return out


def get_email_directory() -> dict[tuple[str, str], dict]:
    data = http_json("GET", f"{MORPHEUS}/api/employees/email-eligible", timeout=30)
    out = {}
    for e in data.get("employees", []):
        out[_name_key(e.get("name", ""))] = e
    return out


def get_ftb_summary(code: str | None) -> dict:
    if not code:
        return {}
    try:
        return http_json("GET", f"{MORPHEUS}/api/employees/{code}/summary?recent_limit=6", timeout=30)
    except Exception as e:
        return {"error": str(e)}


def build_targets(include_ftb: bool = True) -> list[dict]:
    pastors = load_pathways()
    directory = get_email_directory()
    book = load_book_2025()
    targets = []
    for p in pastors:
        name_norm = _norm_name(p.get("name", ""))
        assignment_norm = (p.get("assignment") or "").strip().lower()
        if name_norm in EXCLUDED_ALLOWANCE_NAMES:
            continue
        if assignment_norm in EXCLUDED_ALLOWANCE_ASSIGNMENTS and name_norm not in ALLOWANCE_ELIGIBLE_NAME_EXCEPTIONS:
            continue
        key = _name_key(p["name"])
        emp = directory.get(key) or {}
        # Direct fixes where source spelling differs.
        if not emp and p["name"].lower().startswith("rangi"):
            emp = next((v for v in directory.values() if v.get("code") == "EIAI01"), {})
        ftb = get_ftb_summary(emp.get("code")) if include_ftb and emp else {}
        totals = ftb.get("totals") or {}
        b = book.get(key, {})
        targets.append({
            **p,
            "code": emp.get("code"),
            "email": emp.get("email"),
            "email_name": emp.get("name"),
            "ftb_balance": totals.get("balance"),
            "ftb_ytd_credit": totals.get("ytd_credit"),
            "ftb_ytd_debit": totals.get("ytd_debit"),
            "ftb_as_of": totals.get("as_of"),
            "ftb_error": ftb.get("error"),
            "pde_allowance_2026": 1800.0,
            "evangelist_extra_2026": 900.0 if "evangel" in (p.get("assignment", "") + " " + p.get("churches", "")).lower() else 0.0,
            "first_year_intern_extra_2026": 900.0 if (p.get("snsw_years") in (1, 1.0)) else 0.0,
            **b,
        })
    return targets


def email_html(t: dict, note: str = "") -> str:
    name = escape(t.get("name") or "")
    first = escape((t.get("name") or "there").split()[0])
    rows = "".join([
        f"<tr><td>Employee Exempt Benefits / FTB balance</td><td style='text-align:right'>{_money(t.get('ftb_balance'))}</td><td>MYOB acct 312510 via Benefits Tracker; as at {escape(str(t.get('ftb_as_of') or 'current sync'))}</td></tr>",
        f"<tr><td>2026 Professional Development & Equipment allowance</td><td style='text-align:right'>{_money(t.get('pde_allowance_2026'))}</td><td>2026 policy: base allowance for Ministerial Schedule employees</td></tr>",
        f"<tr><td>2026 Evangelist additional allowance</td><td style='text-align:right'>{_money(t.get('evangelist_extra_2026'))}</td><td>Only where approved/applicable</td></tr>",
        f"<tr><td>2026 First-year intern additional allowance</td><td style='text-align:right'>{_money(t.get('first_year_intern_extra_2026'))}</td><td>Only first-year interns</td></tr>",
    ])
    if t.get("book_balance_2025") is not None:
        rows += f"<tr><td>Book/equipment balance from 2025 workbook</td><td style='text-align:right'>{_money(t.get('book_balance_2025'))}</td><td>Prior workbook reference, not final 2026 MYOB balance</td></tr>"
    note_block = f"<p>{escape(note)}</p>" if note else ""
    return f"""
<div style='font-family:-apple-system,Segoe UI,Arial,sans-serif;color:#1f2937;line-height:1.5;max-width:760px'>
<p>Hi {first},</p>
<p>For Monday's pastoral budget/allowance demo, here is your current allowance snapshot held in the finance dashboard.</p>
<table style='border-collapse:collapse;width:100%;font-size:14px'>
<thead><tr style='background:#f3f4f6'><th style='text-align:left;padding:8px'>Item</th><th style='text-align:right;padding:8px'>Amount</th><th style='text-align:left;padding:8px'>Source / note</th></tr></thead>
<tbody>{rows}</tbody></table>
<p><strong>Appointment context:</strong> {escape(t.get('district') or '')} — {escape(t.get('assignment') or '')}; {escape(t.get('churches') or '')}</p>
{note_block}
<p>If anything looks off, reply to this email and Finance can correct the source data before final release.</p>
<p style='font-size:12px;color:#6b7280'>Sent via snswfinance@adventist.bot from the CFO budget dashboard · {date.today().isoformat()}</p>
</div>"""


def send_emails(payload: dict) -> dict:
    dry_run = bool(payload.get("dry_run", True))
    test_to = payload.get("test_to")
    only_codes = set(payload.get("only_codes") or [])
    note = payload.get("note") or ""
    confirm = payload.get("confirm")
    if not dry_run and not ALLOW_LIVE_SEND:
        raise ValueError("Live send disabled in local CFO dashboard. Set SNSW_ALLOWANCE_EMAIL_LIVE_SEND=1 only after Kyle explicitly approves a live send session.")
    if not dry_run and confirm != CONFIRM_TOKEN:
        raise ValueError(f"Live send blocked. confirm must be {CONFIRM_TOKEN!r}.")
    targets = build_targets(include_ftb=True)
    if only_codes:
        targets = [t for t in targets if t.get("code") in only_codes]
    results = []
    for t in targets:
        to = [test_to] if test_to else ([t.get("email")] if t.get("email") else [])
        subject = payload.get("subject") or f"Your 2026 allowance snapshot — {date.today().isoformat()}"
        body = email_html(t, note=note)
        result = {"name": t.get("name"), "code": t.get("code"), "to": to, "subject": subject, "dry_run": dry_run}
        if not to:
            result.update({"status": "skipped", "reason": "no email matched"})
        elif dry_run:
            result.update({"status": "preview", "html": body})
        else:
            try:
                sent = http_json("POST", f"{MORPHEUS}/api/email/send", {"to": to, "subject": subject, "body": body, "html": True}, timeout=60)
                result.update({"status": "sent", "morpheus": sent})
            except Exception as e:
                result.update({"status": "error", "error": str(e)})
        results.append(result)
    return {"dry_run": dry_run, "count": len(results), "sent": sum(1 for r in results if r.get("status") == "sent"), "results": results}


class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(DASH), **kwargs)

    def _json(self, status: int, data: dict):
        raw = json.dumps(data, indent=2, default=str).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET,POST,OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)

    def do_OPTIONS(self):
        self._json(200, {"ok": True})

    def do_GET(self):
        if self.path.startswith("/api/allowance-emails/preview"):
            try:
                targets = build_targets(include_ftb=True)
                self._json(200, {"generated_at": datetime.now().isoformat(timespec="seconds"), "targets": targets, "sources": {"pathways": str(PATHWAYS), "pde_policy": str(PDE_POLICY), "book_2025": str(BOOK_2025), "morpheus": MORPHEUS}})
            except Exception as e:
                self._json(500, {"error": str(e)})
            return
        super().do_GET()

    def do_POST(self):
        if self.path.startswith("/api/allowance-emails/send"):
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
                self._json(200, send_emails(payload))
            except Exception as e:
                self._json(500, {"error": str(e)})
            return
        self._json(404, {"error": "not found"})


def main():
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8791
    httpd = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"Allowance email dashboard server: http://127.0.0.1:{port}/allowance-email-dashboard.html", flush=True)
    httpd.serve_forever()


if __name__ == "__main__":
    main()
