#!/usr/bin/env python3
from pathlib import Path
import json, html, datetime
from openpyxl import load_workbook
from cfo_stripe_theme import apply_stripe_theme, ensure_theme_file

ROOT=Path('/Users/snswcommunications')
OD=ROOT/'Library/CloudStorage/OneDrive-Seventh-dayAdventistChurch(SouthPacific)'
OUT_DIR=ROOT/'Hermes-CFO/briefings/dashboards'
PAY_DIR=ROOT/'Hermes-CFO/finance/payroll-staff-costs'
OUT_HTML=OUT_DIR/'field-pastoral-staffing-dashboard.html'
OUT_JSON=OUT_DIR/'field-pastoral-staffing-dashboard-data.json'
BUDGET_2026=OD/'Files - SNSW-Finance - Finance/1. SNC/Budget/2026/Budgets 2026.xlsx'
TRIAL=OD/'Files - SNSW-Finance - Finance/1. SNC/Audit/2025/Work papers/GL Trial Balance - Reconciliation migration to MYOB July 25.xlsx'
SOURCE_MAP=ROOT/'Hermes-CFO/projects/cfo-second-brain/indexes/field-pastoral-budget-staffing-sources.md'
STAFF_DASH=OUT_DIR/'staff-cost-dashboard.html'
BUDGET_DASH=OUT_DIR/'department-budget-dashboard.html'
DEPT_DATA=OUT_DIR/'department-budget-dashboard-data.json'
PASTORAL_MAP='http://127.0.0.1:8094/'

def n(v):
    try: return float(v or 0)
    except Exception: return 0.0

def money(v):
    s=f"${abs(n(v)):,.0f}"
    return f"({s})" if n(v)<0 else s

def esc(v): return html.escape('' if v is None else str(v))


def clean_display_text(value):
    text = '' if value is None else str(value)
    replacements = {
        'Cheif Financial Officer': 'Chief Financial Officer',
        'General Secratary': 'General Secretary',
        'Ministerial Secratary': 'Ministerial Secretary',
        'Payrol Clerk': 'Payroll Clerk',
        'Pr. Justin Lawnan': 'Pr. Justin Lawman',
        'Casual/AAV/random - low direct budget impact': 'Unclassified / casual / AAV / low direct budget impact',
        'Kyle role/category overrides': 'manual role/category corrections',
        'Kyle-classified role/category layer': 'manual role/category corrections',
        'source truth': 'source basis',
        'Source truth': 'Source basis',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def budget_2026_field():
    wb=load_workbook(BUDGET_2026,read_only=True,data_only=True,keep_links=False); ws=wb['Budgets']
    headers=[c.value for c in ws[1]]; idx={h:i for i,h in enumerate(headers) if h}
    rows=[]; total=0.0
    staff_accts={'701010','701030','701040','701060','701070','701200','702010','702030','702070','702080','702110','702130','703015','703090','703241','703850'}
    for r in ws.iter_rows(min_row=2,values_only=True):
        acct=str(r[idx['Account']] or '')
        sub=str(r[idx['Subaccount']] or '')
        if not sub.startswith('FLD') or acct not in staff_accts: continue
        amt=abs(n(r[idx['Amount']]))
        desc=str(r[idx['Description']] or '')
        total += amt
        rows.append({'account':acct,'subaccount':sub,'description':desc,'budget':amt})
    return {'total':total,'rows':rows,'source':str(BUDGET_2026)}


def approved_field_budget_control():
    if not DEPT_DATA.exists():
        detail = budget_2026_field()
        return {'total': detail['total'], 'lines': [], 'source': str(DEPT_DATA), 'period': 'approved department budget dashboard data not found', 'status': 'fallback_to_workbook_detail'}
    data = json.loads(DEPT_DATA.read_text(encoding='utf-8'))
    for dept in data.get('departments', []):
        if str(dept.get('name', '')).upper() == 'FIELD':
            return {
                'total': n(dept.get('budget')),
                'spent': n(dept.get('spent')),
                'remaining': n(dept.get('remaining')),
                'lines': dept.get('lines', []),
                'source': str(DEPT_DATA),
                'period': (data.get('period_context') or {}).get('budget_period_label', 'approved budget period not stated'),
                'actual_period': (data.get('period_context') or {}).get('actual_period_label', 'actual period not stated'),
                'status': 'approved_pdf_control_total_from_department_dashboard',
            }
    detail = budget_2026_field()
    return {'total': detail['total'], 'lines': [], 'source': str(DEPT_DATA), 'period': 'FIELD row not found in approved department dashboard data', 'status': 'fallback_to_workbook_detail'}

def saved_staff_analysis():
    p=PAY_DIR/'current_25_26_staff_allocation_with_overrides.csv'
    if not p.exists():
        return {'by_category':[], 'field_people':[], 'direct_conference_total':0, 'source':str(p)}
    import csv
    rows=list(csv.DictReader(p.open()))
    by={}; field=[]; all_people=[]; direct_total=0
    include={'Admin / Executive','Finance','Department director','Department support','Field / pastoral','Other conference'}
    for r in rows:
        cat=r.get('analysis_category') or r.get('final_category') or r.get('category') or 'Unclassified'
        cost=n(r.get('cost_25_26'))
        by.setdefault(cat, {'category':cat,'people':0,'cost':0.0})
        by[cat]['people']+=1; by[cat]['cost']+=cost
        if cat in include: direct_total += cost
        person={'staff_id':r.get('staff_id'),'name':r.get('payroll_name'),'match_name':r.get('match_name') or '', 'match_score':r.get('match_score') or '', 'job_or_area':r.get('job_or_area') or '', 'role':r.get('role') or '', 'notes':r.get('notes') or '', 'category':cat, 'cost_25_26':cost}
        all_people.append(person)
        if cat=='Field / pastoral':
            field.append(person)
    return {'by_category':sorted(by.values(), key=lambda x:x['cost'], reverse=True),'field_people':sorted(field,key=lambda x:x['cost_25_26'],reverse=True),'all_people':sorted(all_people,key=lambda x:x['cost_25_26'],reverse=True),'direct_conference_total':direct_total,'source':str(p)}

def david_funding():
    wb=load_workbook(TRIAL,read_only=True,data_only=True,keep_links=False); ws=wb['Data']
    david=[]; funding=[]
    for row in ws.iter_rows(values_only=True):
        text=' | '.join('' if v is None else str(v) for v in row)
        low=text.lower()
        if 'bogd01' in low or 'bogitini' in low:
            # Account is col 5, description col 6, subaccount col 8, ending balance-ish col 14/15 varies.
            david.append({'account':row[4], 'description':row[5], 'type':row[6], 'subaccount':row[7], 'amount':n(row[13]), 'function':row[32] if len(row)>32 else '', 'raw':text[:900]})
        if 'tr00184' in low or 'dubbo first nations funding' in low:
            funding.append({'account':row[4], 'description':row[5], 'type':row[6], 'subaccount':row[7], 'amount':n(row[13]), 'raw':text[:900]})
    expense=sum(x['amount'] for x in david if str(x['type']).lower()=='expense')
    liability=sum(x['amount'] for x in funding if str(x['type']).lower()=='liability')
    return {'david_expense_rows':david,'funding_rows':funding,'david_expense_total':expense,'funding_liability_balance':liability,'source':str(TRIAL),'status':'probable funding lane; not yet proof of salary offset'}

def main():
    OUT_DIR.mkdir(parents=True,exist_ok=True)
    # Do not hard-code historical trend figures. Older SUN/actual files exist locally, but
    # they have not yet been parsed through a repeatable, verified index in this generator.
    # Until that exists, the dashboard must show an honest not-yet-indexed state.
    actual_trend=[]
    workbook_detail=budget_2026_field()
    approved_budget=approved_field_budget_control()
    data={'generated_at':datetime.datetime.now().isoformat(timespec='seconds'), 'periods':{
        'budget':'FY2026 approved FIELD budget control total from Final budget 2026 PDF via department dashboard; old Budgets 2026.xlsx rows shown only as detail/reconciliation support',
        'payroll':'FY2025-26 payroll allocation extract with manual role/category corrections',
        'history':'Historical FY2015-FY2024 field actuals not yet indexed in a repeatable parser',
        'trial_balance':'August 2025 GL migration trial balance evidence only'
    }, 'budget_2026':approved_budget, 'budget_workbook_detail': workbook_detail, 'budget_reconciliation': {'approved_control_total': approved_budget.get('total'), 'workbook_detail_total': workbook_detail.get('total'), 'difference': n(approved_budget.get('total')) - n(workbook_detail.get('total')), 'approved_source': approved_budget.get('source'), 'detail_source': workbook_detail.get('source'), 'status': 'approved PDF control total is authoritative; workbook rows are detail support until reconciled'}, 'staff':saved_staff_analysis(), 'historical_actual_trend':actual_trend, 'history_status':{
        'status':'not yet indexed',
        'confidence':'No chart shown until historical workbook/SUN parsing is repeatable and source rows are attributed.',
        'next_action':'Build a local historical field-staff-cost index with source workbook, account, function, FY, actual, and confidence columns.'
    }, 'david':david_funding(), 'sources':[
        {'label':'Source map','path':str(SOURCE_MAP),'use':'Routing and source attribution'},
        {'label':'2026 approved FIELD budget control','path':str(DEPT_DATA),'use':'Authoritative approved FY2026 FIELD budget control total from Final budget 2026 PDF via department dashboard'},
        {'label':'Old budget workbook FLD detail','path':str(BUDGET_2026),'use':'Detail/reconciliation support only; not the displayed control total'},
        {'label':'Staff role overrides','path':str(PAY_DIR/'staff-role-overrides.json'),'use':'manual role/category corrections'},
        {'label':'Staff payroll allocation extract','path':str(PAY_DIR/'current_25_26_staff_allocation_with_overrides.csv'),'use':'FY2025-26 payroll people/cost by role category'},
        {'label':'Historical field actuals','path':'Local historical SUN/workbook sources under OneDrive-derived finance folders','use':'Not yet indexed; trend intentionally suppressed until repeatable parsing is verified'},
        {'label':'David / First Nations trial balance evidence','path':str(TRIAL),'use':'BOGD01 expense and TR00184 Dubbo First Nations Funding liability lane'}
    ]}
    OUT_JSON.write_text(json.dumps(data,indent=2),encoding='utf-8')
    chart=json.dumps(data)
    field_rows=''.join(f"<tr><td>{esc(clean_display_text(x['name']))}</td><td>{esc(clean_display_text(x.get('role') or 'Field / pastoral'))}</td><td>{money(x['cost_25_26'])}</td><td>{esc(clean_display_text(x.get('notes','')))}</td></tr>" for x in data['staff']['field_people'])
    cat_rows=''.join(f"<tr><td>{esc(clean_display_text(x['category']))}</td><td>{x['people']}</td><td>{money(x['cost'])}</td></tr>" for x in data['staff']['by_category'])
    category_options=''.join(f"<option value=\"{esc(clean_display_text(x['category']))}\">{esc(clean_display_text(x['category']))} — {x['people']} people — {money(x['cost'])}</option>" for x in data['staff']['by_category'])
    all_people_rows=''.join(f"<tr data-cat=\"{esc(clean_display_text(x['category']))}\"><td><b>{esc(clean_display_text(x['name']))}</b><br><span class='mini'>{esc(x['staff_id'])}</span></td><td>{esc(clean_display_text(x['category']))}</td><td>{money(x['cost_25_26'])}</td><td>{esc(clean_display_text(x.get('job_or_area') or x.get('role') or ''))}</td><td>{esc(clean_display_text(x.get('match_name') or ''))}<br><span class='mini'>{esc(clean_display_text(x.get('notes') or ''))}</span></td></tr>" for x in data['staff']['all_people'])
    budget_rows=''.join(f"<tr><td>{esc(x.get('account',''))}</td><td>{esc(x.get('description') or x.get('line',''))}</td><td>{money(x.get('budget'))}</td></tr>" for x in (data.get('budget_workbook_detail',{}).get('rows') or data['budget_2026'].get('lines', [])))
    sources=''.join(f"<tr><td>{esc(s['label'])}</td><td>{esc(s['use'])}</td><td class='src'>{esc(s['path'])}</td></tr>" for s in data['sources'])
    david_rows=''.join(f"<tr><td>{esc(x['account'])}</td><td>{esc(x['description'])}</td><td>{esc(x['type'])}</td><td>{money(x['amount'])}</td></tr>" for x in data['david']['david_expense_rows'][:20])
    fund_rows=''.join(f"<tr><td>{esc(x['account'])}</td><td>{esc(x['description'])}</td><td>{esc(x['type'])}</td><td>{money(x['amount'])}</td></tr>" for x in data['david']['funding_rows'])
    recon=data.get('budget_reconciliation',{})
    recon_panel=f"<div class='card span12 warn'><div class='label'>Budget source reconciliation</div><div class='note'><b>Approved FIELD control total:</b> {money(recon.get('approved_control_total'))}. <b>Old workbook detail rows:</b> {money(recon.get('workbook_detail_total'))}. <b>Difference:</b> {money(recon.get('difference'))}. Use the approved control total for board/CFO discussion; use workbook rows only for line-detail support until fully reconciled.</div></div>"
    history_status=data.get('history_status',{})
    history_trend=data.get('historical_actual_trend') or []
    history_change='Not indexed'
    if len(history_trend) >= 2 and history_trend[0].get('actual'):
        history_change=f"{((history_trend[-1]['actual']/history_trend[0]['actual']-1)*100):.0f}%"
    history_panel = "<canvas id='trend'></canvas>" if history_trend else f"<div class='notice'><b>History not yet indexed.</b><br>{esc(history_status.get('confidence','Historical actuals are not yet verified.'))}<br><span class='mini'>Next: {esc(history_status.get('next_action','Build verified historical index.'))}</span></div>"
    html_doc=f"""<!doctype html><html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>Field / Pastoral Staffing Budget Dashboard</title><style>
.answer-strip{{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:14px;margin-top:16px}}.answer-strip .card{{min-height:132px}}.source-chip-row{{display:flex;flex-wrap:wrap;gap:8px;margin-top:10px}}.source-chip-row span{{border:1px solid #e5edf5;background:#fff;border-radius:4px;padding:6px 8px;color:#64748d;font-size:12px}}details.compact{{margin-top:14px}}details.compact summary{{cursor:pointer;color:#533afd;font-weight:500}}@media(max-width:900px){{.answer-strip{{grid-template-columns:repeat(2,minmax(0,1fr))}}}}@media(max-width:620px){{.answer-strip{{grid-template-columns:1fr}}}}
</style></head><body><div class='wrap'><h1>Field / Pastoral Staffing Budget Dashboard</h1><div class='sub'>Justin Lawman decision support: approved FIELD budget, exact pastoral payroll rows, history status, and funding-offset evidence.</div><div class='answer-strip'><div class='card'><div class='label'>Answer</div><div class='note'><b>Use the top figures first.</b> Budget is authority; payroll is named people cost; history is not safe yet.</div></div><div class='card'><div class='label'>Risk</div><div class='note'>Do not use prior-year Field trend until SUN/workbook history is indexed with source rows.</div></div><div class='card'><div class='label'>Source</div><div class='source-chip-row'><span>Budget: approved FY2026 FIELD control</span><span>Payroll: FY2025-26 allocation extract</span><span>Funding: Aug 2025 GL migration evidence</span></div></div></div><details class='compact card'><summary>Basis and caveats</summary><div class='note'><b>Budget:</b> {esc(data['periods']['budget'])}. <b>Payroll:</b> {esc(data['periods']['payroll'])}. <b>History:</b> {esc(data['periods']['history'])}. Field budget and field payroll/person cost are related but not identical.</div></details><div class='grid'>
<div class='card span3'><div class='label'>2026 approved FIELD budget</div><div class='value'>{money(data['budget_2026']['total'])}</div><div class='sub'>{esc(data['periods']['budget'])}</div></div>
<div class='card span3'><div class='label'>FY2025-26 field payroll mapped</div><div class='value'>{money(sum(x['cost_25_26'] for x in data['staff']['field_people']))}</div><div class='sub'>{len(data['staff']['field_people'])} field/pastoral people; {esc(data['periods']['payroll'])}</div></div>
<div class='card span3'><div class='label'>Historical field staff-cost trend</div><div class='value'>{esc(history_change)}</div><div class='sub'>{esc(data['periods']['history'])}</div></div>
<div class='card span3 warn'><div class='label'>David funding lane found</div><div class='value'>{money(data['david']['funding_liability_balance'])}</div><div class='sub'>TR00184 Dubbo First Nations Funding — {esc(data['periods']['trial_balance'])}</div></div>
{recon_panel}
<div class='card span12'><div class='label'>Historical field actual staff-cost trend</div>{history_panel}<div class='note'>Historical actual spend trend basis: {esc(data['periods']['history'])}. This page stays focused on field/pastoral staffing evidence and will not show historical figures until source parsing is verified.</div></div>
<div class='card span6'><div class='label'>FY2025-26 staff cost by category</div><table><thead><tr><th>Category</th><th>People</th><th>FY2025-26 payroll cost</th></tr></thead><tbody>{cat_rows}</tbody></table><div class='note'>Use the staff detail table below for exact names and FY2025-26 figures — this summary is only a roll-up.</div></div>
<div class='card span6'><div class='label'>Old workbook FLD detail rows</div><div class='note'>These rows come from Budgets 2026.xlsx and do not equal the approved PDF control total above. Treat as detail/reconciliation support, not authority.</div><table><thead><tr><th>Acct</th><th>Description</th><th>Old workbook FLD amount</th></tr></thead><tbody>{budget_rows}</tbody></table></div>
<div class='card span12' id='staff-detail'><div class='label'>Staff cost detail by person</div><div class='note'><b>Fast path:</b> choose a category, scan exact names/costs, then use the pastoral map only for assignment/location context. Payroll cost is the source lane for dollars.</div><div style='max-width:520px;margin:12px 0'><select id='catFilter' onchange='filterStaffCategory()'><option value=''>All categories</option>{category_options}</select></div><table id='staffDetail'><thead><tr><th>Name</th><th>Category</th><th>FY2025-26 payroll cost</th><th>Role / area</th><th>Matched source / notes</th></tr></thead><tbody>{all_people_rows}</tbody></table></div>
<div class='card span12' id='field-people'><div class='label'>Field / pastoral people mapped in FY2025-26 payroll extract</div><table><thead><tr><th>Name</th><th>Role</th><th>FY2025-26 payroll cost</th><th>Notes</th></tr></thead><tbody>{field_rows}</tbody></table></div>
<div class='card span6 warn'><div class='label'>David Bogitini — posted field expense evidence</div><div class='note'>Status: {esc(data['david']['status'])}. Expense rows below are from the 08-2025 GL migration trial balance; this supports field coding but is not yet the funding agreement.</div><table><thead><tr><th>Acct</th><th>Description</th><th>Type</th><th>Aug 2025 GL migration amount</th></tr></thead><tbody>{david_rows}</tbody></table></div>
<div class='card span6 warn'><div class='label'>David / First Nations funding evidence found</div><table><thead><tr><th>Acct</th><th>Description</th><th>Type</th><th>Aug 2025 GL migration balance</th></tr></thead><tbody>{fund_rows}</tbody></table><div class='note'>Next evidence required before treating as a confirmed offset: grant/funding letter, bank receipt, trust ledger, journal memo, or correspondence linking this trust to David’s role.</div></div>
<div class='card span12'><div class='label'>Source attribution</div><table><thead><tr><th>Source</th><th>Use</th><th>Path</th></tr></thead><tbody>{sources}</tbody></table></div>
</div></div><script>const D={chart};function money(x){{return '$'+Math.round(x).toLocaleString()}}function filterStaffCategory(){{const v=document.getElementById('catFilter').value;document.querySelectorAll('#staffDetail tbody tr').forEach(tr=>tr.style.display=(!v||tr.dataset.cat===v)?'':'none')}}function draw(){{let c=document.getElementById('trend'); if(!c || !D.historical_actual_trend || D.historical_actual_trend.length<2) return; let ctx=c.getContext('2d');c.width=c.clientWidth*2;c.height=320*2;ctx.scale(2,2);let W=c.clientWidth,H=320,max=Math.max(...D.historical_actual_trend.map(x=>x.actual))*1.12,min=Math.min(...D.historical_actual_trend.map(x=>x.actual))*0.90;ctx.clearRect(0,0,W,H);ctx.font='13px Source Sans 3, system-ui';ctx.textBaseline='middle';ctx.strokeStyle='#38bdf8';ctx.lineWidth=4;ctx.beginPath();D.historical_actual_trend.forEach((d,i)=>{{let x=80+i*(W-150)/(D.historical_actual_trend.length-1),y=H-58-(d.actual-min)/(max-min)*(H-120);if(i)ctx.lineTo(x,y);else ctx.moveTo(x,y);}});ctx.stroke();D.historical_actual_trend.forEach((d,i)=>{{let x=80+i*(W-150)/(D.historical_actual_trend.length-1),y=H-58-(d.actual-min)/(max-min)*(H-120);ctx.fillStyle='#38bdf8';ctx.beginPath();ctx.arc(x,y,3,0,Math.PI*2);ctx.fill();ctx.fillStyle='#273951';ctx.fillText(d.year,x-16,H-24);ctx.fillStyle='#061b31';ctx.font='600 13px Source Sans 3, system-ui';ctx.fillText(money(d.actual),x-38,y-18);ctx.font='13px Source Sans 3, system-ui';}});}}draw();</script></body></html>"""
    ensure_theme_file(OUT_DIR)
    OUT_HTML.write_text(apply_stripe_theme(html_doc),encoding='utf-8')
    print(OUT_HTML)
if __name__=='__main__': main()
